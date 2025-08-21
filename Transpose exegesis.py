
import tkinter as tk
# 导入必要的库
# from tkinter import filedialog, ttk: 从 tkinter 模块导入 filedialog（用于文件选择对话框）和 ttk（主题化 Tkinter 部件）
from tkinter import filedialog, ttk
# from tkinter import messagebox: 从 tkinter 模块导入 messagebox（用于显示消息框）
from tkinter import messagebox
# import pandas as pd: 导入 pandas 库，通常用于数据处理和分析，并约定别名为 pd
import pandas as pd
# import numpy as np: 导入 numpy 库，通常用于科学计算，特别是数组操作，并约定别名为 np
import numpy as np
# from openpyxl.workbook import Workbook: 从 openpyxl 库导入 Workbook 类，用于创建新的 Excel 工作簿
from openpyxl.workbook import Workbook
# 从 openpyxl.utils 模块导入 get_column_letter 函数，用于将列索引转换为 Excel 列字母（例如，1 -> A, 2 -> B）
from openpyxl.utils import get_column_letter
# import os: 导入 os 模块，用于与操作系统进行交互，例如文件路径操作
import os
# from tkinterdnd2 import TkinterDnD, DND_FILES: 从 tkinterdnd2 库导入 TkinterDnD（支持拖放功能的 Tkinter 根窗口）和 DND_FILES（拖放文件类型常量）
from tkinterdnd2 import TkinterDnD, DND_FILES

# import time # 这是一个被注释掉的导入语句，表示 time 模块在此代码中未使用


# 定义处理 Excel 文件的函数
# file_path: 传入的 Excel 文件路径参数
def process_excel_file(file_path):
    # 检查文件路径是否存在
    if file_path:
        try:
            # 使用 pandas 读取 Excel 文件到 DataFrame
            # df = pd.read_excel(file_path): 读取指定路径的 Excel 文件，并将其内容存储在 DataFrame 对象 df 中
            df = pd.read_excel(file_path)

            # 对 DataFrame 进行转置操作
            # df_transpose = df.transpose(): 将 df 进行行列转置，得到新的 DataFrame df_transpose
            df_transpose = df.transpose()
            # df_index = df_transpose.index: 获取转置后 DataFrame 的索引（即原始 DataFrame 的列名），这些将作为新 Excel 的第一列内容
            df_index = df_transpose.index

            # 将原始 DataFrame 转换为 NumPy 数组
            # arr = np.array(df): 将 df 转换为 NumPy 数组 arr，方便后续按行或列访问数据
            arr = np.array(df)

            # 创建一个新的 Excel 工作簿和工作表
            # wb = Workbook(): 创建一个新的 Excel 工作簿对象
            wb = Workbook()
            # sheet = wb.active: 获取当前活动的工作表，所有数据将写入此工作表
            sheet = wb.active
            # 遍历 arr 数组的每一行（原始 Excel 的每一行）
            # j: 当前行的索引
            for j in range(0, len(arr)):
                # 设置新 Excel 文件中列的宽度
                # sheet.column_dimensions[get_column_letter(j*2+1)].width = 15: 设置奇数列的宽度为 15，用于存放原始 Excel 的列名
                sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
                # 遍历 df_index（原始 Excel 的列名）
                # i: 当前列名的索引
                for i in range(0, len(df_index)):
                    # 将原始 Excel 的列名写入新 Excel 的奇数列
                    # sheet.cell(row=i+1, column=j*2+1).value = df_index[i]: 将 df_index 中的值写入新 Excel 的指定单元格
                    sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
            # 遍历 arr 数组的每一行（原始 Excel 的每一行）
            # k: 当前行的索引
            for k in range(0, len(arr)):
                # 设置新 Excel 文件中列的宽度
                # sheet.column_dimensions[get_column_letter(k*2+2)].width = 15: 设置偶数列的宽度为 15，用于存放原始 Excel 的数据
                sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
                # 遍历 df_index（原始 Excel 的列名）
                # col_idx: 当前列名的索引
                for col_idx in range(0, len(df_index)):
                    # 将原始 Excel 的数据写入新 Excel 的偶数列
                    # sheet.cell(row=col_idx+1, column=k*2+2).value = arr[k][col_idx]: 将 arr 中的值写入新 Excel 的指定单元格
                    sheet.cell(row=col_idx+1, column=k*2+2).value = arr[k][col_idx]

            # 构建输出文件路径并保存
            # file_out = file_path.replace('.xlsx', ''): 移除原始文件路径中的 .xlsx 扩展名
            file_out = file_path.replace('.xlsx', '')
            # file_out = file_out.replace('.xls', ''): 移除原始文件路径中的 .xls 扩展名
            file_out = file_out.replace('.xls', '')
            # wb.save(file_out+'_'+'转置.xlsx'): 将工作簿保存为新的 Excel 文件，文件名在原始文件名的基础上添加 '_转置.xlsx'
            wb.save(file_out+'_'+'转置.xlsx')
            # 显示成功消息框，包含生成文件的路径
            # messagebox.showinfo(message=f"转置成功！！\n生成文件路径：{file_out}_转置.xlsx"): 弹出信息框，告知用户转置成功，并显示生成文件的完整路径
            messagebox.showinfo(message=f"转置成功！！\n生成文件路径：{file_out}_转置.xlsx")
            # 关闭主窗口
            # root.destroy(): 销毁 Tkinter 主窗口，结束应用程序
            root.destroy()
        # 捕获处理文件时可能发生的异常
        except Exception as e:
            # 显示错误消息框
            # messagebox.showerror("错误", f"处理文件时发生错误: {e}"): 弹出错误框，显示错误信息
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")


# 定义选择文件的函数
def select_file():
    # 弹出文件选择对话框
    # filedialog.askopenfilename: 打开一个文件选择对话框，让用户选择文件
    # title="选择Excel文件": 对话框的标题
    # initialdir=os.path.expanduser('~'): 对话框的初始目录设置为用户的主目录
    # filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")]: 允许选择的文件类型，这里是 Excel 文件和所有文件
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        initialdir=os.path.expanduser('~'),
        filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
    # 调用处理 Excel 文件的函数
    process_excel_file(file_path)


# 定义拖放文件的函数
# event: 拖放事件对象，包含了拖放的数据
def drop_file(event):
    # 获取拖放的文件路径
    # event.data: 拖放的数据，通常是文件路径字符串
    file_path = event.data
    # 检查文件路径是否被大括号包围（在某些拖放操作中可能会出现）
    if file_path.startswith('{') and file_path.endswith('}'):
        # 移除大括号
        file_path = file_path[1:-1]
    # 调用处理 Excel 文件的函数
    process_excel_file(file_path)


# 创建 Tkinter 主窗口
# root = TkinterDnD.Tk(): 创建一个支持拖放功能的 Tkinter 根窗口
root = TkinterDnD.Tk()
# 设置窗口标题
# root.title("Excel转置工具"): 设置窗口的标题栏文本
root.title("Excel转置工具")
# 获取屏幕的宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 设置窗口的初始宽度和高度
window_width = 500
window_height = 300

# 计算窗口的x, y坐标，使其居中
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# 设置窗口的几何尺寸和位置
# root.geometry(f'{window_width}x{window_height}+{x}+{y}'): 设置窗口的宽度、高度以及在屏幕上的初始位置
root.geometry(f'{window_width}x{window_height}+{x}+{y}')
# 设置窗口是否可调整大小
# root.resizable(True, True): 允许窗口在水平和垂直方向上调整大小
root.resizable(True, True)

# 配置窗口背景颜色
# root.configure(bg='#f0f0f0'): 设置窗口的背景颜色为浅灰色
root.configure(bg='#f0f0f0')

# 创建主框架
# main_frame = ttk.Frame(root, padding=20): 创建一个 ttk 框架，作为主窗口的子组件，并设置内边距
main_frame = ttk.Frame(root, padding=20)
# 将主框架放置在窗口中
# main_frame.pack(fill='both', expand=True): 将框架打包到窗口中，使其填充可用空间并随窗口大小扩展
main_frame.pack(fill='both', expand=True)

# 配置窗口背景颜色 (重复，可以移除一个)
root.configure(bg='#f0f0f0')

# 创建标题标签
# title_label = ttk.Label(...): 创建一个 ttk 标签，用于显示标题文本
# text="Excel转置工具": 标签显示的文本
# font=('微软雅黑', 16, 'bold'): 设置字体为微软雅黑，大小16，加粗
# foreground='#333333': 设置文本颜色为深灰色
title_label = ttk.Label(main_frame,
                        text="Excel转置工具",
                        font=('微软雅黑', 16, 'bold'),
                        foreground='#333333')
# 将标题标签放置在主框架中
# title_label.pack(pady=(0, 20)): 将标签打包到框架中，设置垂直方向的外部填充
title_label.pack(pady=(0, 20))

# 创建提示文本标签
# text_label = ttk.Label(...): 创建一个 ttk 标签，用于显示操作提示文本
# text="请选择要转置的Excel文件或将文件拖拽到此处": 标签显示的文本
# font=('微软雅黑', 12): 设置字体为微软雅黑，大小12
text_label = ttk.Label(main_frame,
                       text="请选择要转置的Excel文件或将文件拖拽到此处",
                       font=('微软雅黑', 12))
# 将提示文本标签放置在主框架中
# text_label.pack(pady=(0, 30)): 将标签打包到框架中，设置垂直方向的外部填充
text_label.pack(pady=(0, 30))

# 配置按钮样式
# style = ttk.Style(): 创建一个 ttk 样式对象
style = ttk.Style()
# style.configure('TButton', font=('微软雅黑', 12), padding=10): 配置名为 'TButton' 的样式，设置字体和内边距
style.configure('TButton', font=('微软雅黑', 12), padding=10)

# 创建选择文件按钮
# btn = ttk.Button(...): 创建一个 ttk 按钮
# text="选择文件": 按钮显示的文本
# command=select_file: 按钮点击时调用的函数
# style='TButton': 应用之前定义的 'TButton' 样式
btn = ttk.Button(main_frame,
                 text="选择文件",
                 command=select_file,
                 style='TButton')
# 将按钮放置在主框架中
# btn.pack(): 将按钮打包到框架中
btn.pack()

# 创建底部提示标签
# footer_label = ttk.Label(...): 创建一个 ttk 标签，用于显示底部提示文本
# text="支持.xls和.xlsx格式": 标签显示的文本
# font=('微软雅黑', 9): 设置字体为微软雅黑，大小9
# foreground='#666666': 设置文本颜色为灰色
footer_label = ttk.Label(main_frame,
                         text="支持.xls和.xlsx格式",
                         font=('微软雅黑', 9),
                         foreground='#666666')
# 将底部提示标签放置在主框架中
# footer_label.pack(side=tk.BOTTOM, pady=(20, 0)): 将标签打包到框架底部，设置垂直方向的外部填充
footer_label.pack(side=tk.BOTTOM, pady=(20, 0))

# 注册拖放目标
# root.drop_target_register(DND_FILES): 注册主窗口为拖放目标，接受文件拖放
root.drop_target_register(DND_FILES)
# 绑定拖放事件
# root.dnd_bind('<<Drop>>', drop_file): 将拖放事件绑定到 drop_file 函数，当文件被拖放到窗口时调用该函数
root.dnd_bind('<<Drop>>', drop_file)

# 启动 Tkinter 事件循环
# root.mainloop(): 启动 Tkinter 的事件循环，使窗口保持显示并响应用户交互，直到窗口关闭
root.mainloop()
