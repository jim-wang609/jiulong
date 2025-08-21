import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import qrcode
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage

row_max = 152
page_max = 10
page_column_max = 3
page_zoom = 75
filepath = '4.xlsx'
sheet_name = '8成品二算料单标准'
df = pd.read_excel(filepath, sheet_name=sheet_name)
# wb = openpyxl.load_workbook(filepath)
# ws = wb.create_sheet(title='生成二维码')
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = df.columns[7]
ws.cell(row=2, column=1).value = df.columns[16]
new_i, new_j, i_count, j_count = 3, 1, 0, 0


def generate_qrcode_image(data, size):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    transparent_bg = True
    if transparent_bg:
        datas = img.getdata()
        new_data = []
        for item in datas:
            if item[0] == 255 and item[1] == 255 and item[2] == 255:
                new_data.append((255, 255, 255, 0))
            else:
                new_data.append(item)
        img.putdata(new_data)

    img = img.resize((size, size), PILImage.Resampling.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return Image(buffer)


for i in range(2, row_max):
    if pd.isna(df.iloc[i, 7]) or df.iloc[i, 7] == " ":
        pass
    else:
        ws.cell(row=new_i + page_max * i_count, column=new_j + 2 * j_count).value = df.iloc[i, 7]
        qr_img = generate_qrcode_image(df.iloc[i, 7], 120)
        target_cell = ws.cell(row=new_i + page_max * i_count, column=new_j + 2 * j_count + 1)
        qr_img.anchor = f"{target_cell.column_letter}{target_cell.row}"
        ws.add_image(qr_img)
        new_i += 1
        if new_i == page_max + 3:
            new_i = 3
            j_count += 1
        if j_count == 3:
            j_count = 0
            i_count += 1
for i in range(3):
    ws.column_dimensions[get_column_letter(1 + 2 * i)].width = 29.22
    ws.column_dimensions[get_column_letter(2 + 2 * i)].width = 16
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=1, max_row=2):
    for cell_i in row:
        cell_i.alignment = Alignment(horizontal='center', vertical='center')
        cell_i.border = border
        cell_i.font = Font(size=18)
for row in ws.iter_rows(min_row=3):
    for cell_i in row:
        cell_i.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        cell_i.border = border
        cell_i.font = Font(size=20)
for row in range(1, 3):
    ws.row_dimensions[row].height = 22.5
for row in range(3, ws.max_row + 1):
    ws.row_dimensions[row].height = 98
ws.page_margins.left = 0
ws.page_margins.right = 0
ws.page_margins.top = 0
ws.page_margins.bottom = 0
ws.print_title_rows = '1:2'
ws.print_area = f'A1:F{ws.max_row}'
ws.page_setup.scale = page_zoom
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.horizontalCentered = True
ws.page_setup.verticalCentered = True
ws.merge_cells('A1:F1')
ws.merge_cells('A2:F2')
wb.save('output.xlsx')
os.startfile('output.xlsx')
# wb.save(filepath)
# os.startfile(filepath)
