import tkinter as tk
from tkinter import filedialog  # 导入文件选择模块
from tkinter import messagebox
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
# import time


def select_file():
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择文件",  # 对话框标题
        filetypes=[("表格", "*.xls"), ("表格", "*.xlsx*")]  # 可选择的文件类型
    )
    if file_path:  # 如果选择了文件
        file_in = file_path
        df = pd.read_excel(file_in)

        df_transpose = df.transpose()
        df_index = df_transpose.index

        arr = np.array(df)
        # arr_transpose = arr.transpose()
        #print(arr)
        # print(arr_transpose)
        # print(df_transpose.columns)
        # print(df_values)
        # print(df_index)
        # print(df.values)
        wb = Workbook()
        sheet = wb.active
        for j in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
            for i in range(0, len(df_index)):
                sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
        for k in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
            for l in range(0, len(df_index)):
                sheet.cell(row=l+1, column=k*2+2).value = arr[k][l]

        file_out = file_in.replace('.xlsx', '')
        file_out = file_out.replace('.xls', '')
        wb.save(file_out+'_'+'转置.xlsx')
        messagebox.showinfo(message="转置成功！！")
        root.destroy()

# 创建窗口和按钮
root = tk.Tk()

root.title("excel转置脚本")
#root.iconbitmap("图标.ico")
root.geometry('500x250+500+300')
btn = tk.Button(root, text="选择文件", command=select_file)
text_label = tk.Label(root, text="请选择要转置的文件", font="宋体 14 bold", anchor="e")
text_label.pack(padx=5, pady=10)
btn.pack(anchor=tk.CENTER)
root.mainloop()
