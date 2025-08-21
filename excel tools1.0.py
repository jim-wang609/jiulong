import tkinter as tk
# 导入必要的库
# filedialog: 用于打开文件选择对话框
# ttk: Tkinter主题化部件，提供更现代的UI组件
# messagebox: 用于显示消息框，如信息、警告和错误
from tkinter import filedialog, ttk, messagebox
# pandas: 用于数据处理和分析，特别是Excel文件的读写
import pandas as pd
# numpy: 用于数值计算，特别是数组操作
import numpy as np
# openpyxl: 用于读写.xlsx文件
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
# os: 用于操作系统相关操作，如路径处理
import os
# tkinterdnd2: 扩展Tkinter，支持拖放功能
from tkinterdnd2 import TkinterDnD, DND_FILES


def process_excel_file(file_path):
    """
    处理Excel文件，将其内容进行转置并保存为新文件。

    Args:
        file_path (str): 待处理的Excel文件的完整路径。
    """
    if file_path:
        try:
            # 读取Excel文件到DataFrame
            df = pd.read_excel(file_path)
            # 对DataFrame进行转置
            df_transpose = df.transpose()
            # 获取转置后DataFrame的索引（原DataFrame的列名）
            df_index = df_transpose.index
            # 将原DataFrame转换为NumPy数组
            arr = np.array(df)

            # 创建一个新的Excel工作簿
            wb = Workbook()
            # 获取当前活动的工作表
            sheet = wb.active

            # 遍历原数据的每一行，将原列名写入新工作表
            for j in range(0, len(arr)):
                # 设置列宽
                sheet.column_dimensions[get_column_letter(j * 2 + 1)].width = 15
                for i in range(0, len(df_index)):
                    # 写入原列名
                    sheet.cell(row=i + 1, column=j * 2 + 1).value = df_index[i]

            # 遍历原数据的每一行，将原数据写入新工作表
            for k in range(0, len(arr)):
                # 设置列宽
                sheet.column_dimensions[get_column_letter(k * 2 + 2)].width = 15
                for col_idx in range(0, len(df_index)):
                    # 写入原数据
                    sheet.cell(row=col_idx + 1, column=k * 2 + 2).value = arr[k][col_idx]

            # 构造输出文件路径
            file_out = file_path.replace('.xlsx', '')
            file_out = file_out.replace('.xls', '')
            # 保存转置后的文件
            wb.save(file_out + '_' + '转置.xlsx')
            # 显示成功消息
            messagebox.showinfo(message=f"转置成功！！\n新文件路径：{file_out}_转置.xlsx")
        except Exception as e:
            # 捕获并显示处理文件时发生的错误
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")

def create_transpose_tool_page(parent_frame):
    """
    创建Excel转置工具页面。

    Args:
        parent_frame (ttk.Frame): 父级框架，用于放置页面组件。
    """
    # 清空父级框架中的所有组件
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def select_file():
        """
        通过文件选择对话框选择Excel文件并进行处理。
        """
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=os.path.expanduser('~'),
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        process_excel_file(file_path)

    def drop_file(event):
        """
        处理拖放到窗口的Excel文件。

        Args:
            event: 拖放事件对象，包含拖放的文件路径。
        """
        file_path = event.data
        # 移除文件路径字符串可能包含的花括号
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        process_excel_file(file_path)

    # 页面标题标签
    title_label = ttk.Label(parent_frame,
                            text="Excel转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    title_label.pack(pady=(0, 20))

    # 提示文本标签
    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))

    # 配置按钮样式
    style = ttk.Style()
    style.configure('TButton', font=('微软雅黑', 12), padding=10)

    # 选择文件按钮
    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()

    # 页脚提示标签
    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    footer_label.pack(side='bottom', pady=(20, 0))

    # 注册拖放目标并绑定拖放事件
    parent_frame.drop_target_register(DND_FILES)
    parent_frame.dnd_bind('<<Drop>>', drop_file)

# 初始化TkinterDnD主窗口
root = TkinterDnD.Tk()
# 设置窗口标题
root.title("Excel工具")

# 获取屏幕的宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 设置窗口的宽度和高度
window_width = 500
window_height = 300

# 计算窗口居中显示的位置
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# 设置窗口大小和位置
root.geometry(f'{window_width}x{window_height}+{x}+{y}')
# 允许窗口大小调整
root.resizable(True, True)
# 设置窗口背景颜色
root.configure(bg='#f0f0f0')

# 创建一个主框架用于容纳所有页面，并设置内边距
main_app_frame = ttk.Frame(root, padding=20)
# 将主框架填充整个窗口并随窗口大小扩展
main_app_frame.pack(fill='both', expand=True)


def show_home_page():
    """
    显示主页内容，包括标题和功能按钮。
    """
    # 清空主应用框架中的所有组件
    for widget in main_app_frame.winfo_children():
        widget.destroy()

    # 主页标题标签
    home_title_label = ttk.Label(main_app_frame,
                                 text="Excel工具",
                                 font=('微软雅黑', 18, 'bold'),
                                 foreground='#333333')
    home_title_label.pack(pady=(20, 30))

    # Excel转置工具按钮，点击后切换到转置工具页面
    transpose_btn = ttk.Button(main_app_frame,
                               text="Excel 转置工具",
                               command=lambda: create_transpose_tool_page(main_app_frame),
                               style='TButton')
    transpose_btn.pack(pady=10)

    

def create_transpose_tool_page(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def select_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=os.path.expanduser('~'),
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        process_excel_file(file_path)

    def drop_file(event):
        file_path = event.data
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        process_excel_file(file_path)

    title_label = ttk.Label(parent_frame,
                            text="Excel转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    title_label.pack(pady=(0, 20))

    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))

    style = ttk.Style()
    style.configure('TButton', font=('微软雅黑', 12), padding=10)

    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()

    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    footer_label.pack(side='bottom', pady=(20, 0))

    parent_frame.drop_target_register(DND_FILES)
    parent_frame.dnd_bind('<<Drop>>', drop_file)

    # 返回主页按钮
    back_button = ttk.Button(parent_frame,
                             text="返回主页",
                             command=show_home_page,
                             style='TButton')
    back_button.pack(pady=10)

# 首次启动时显示主页
show_home_page()
# 启动Tkinter事件循环
root.mainloop()
