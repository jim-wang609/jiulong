import os
import tkinter as tk
from tkinter import ttk  # 导入ttk模块
from tkinter import filedialog  # 导入文件选择模块
from tkinter import messagebox
import math
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tkinterdnd2 import TkinterDnD, DND_FILES
from openpyxl.styles import Font

# 全局变量
file_path_1 = None
sheet = None
entry = None
path_label = None
options = None


# """将表格列字母（如'A'、'AB'）转换为数字（如1、28）"""
def column_letter_to_number(letter):
    number = 0
    for char in letter.upper():
        number = number * 26 + (ord(char) - ord('A') + 1)
    return number


# 拆分单元格
def unmerged(file_path, sheet):
    if file_path:
        try:
            book = load_workbook(file_path)
            ws = book[sheet]
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell_value = ws.cell(row=min_row, column=min_col).value
                ws.unmerge_cells(str(merged_range))
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        ws.cell(row=row_idx, column=col_idx).value = top_left_cell_value
            output_file_path = file_path.replace('.xlsx', '') + '_' + "unmerged.xlsx"
            book.save(output_file_path)
            return output_file_path
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")


# 参数主函数
def execute():
    global file_path_2, sheet
    parameter(file_path_2, sheet, file_path_1)


def parameter(file, sheet, file_2):
    pf = pd.read_excel(file, sheet_name=sheet)
    pf1 = pd.read_excel(file_2, sheet_name=sheet)
    try:
        pf_disposition = pd.read_excel('disposition.xlsx', sheet_name="Sheet1")
        pf_disposition_1 = pd.read_excel('disposition.xlsx', sheet_name="Sheet2")
    except Exception as e:
        messagebox.showinfo(message=f'{e}')
    arr = np.array(pf)
    arr1 = np.array(pf1)
    arr_disposition = np.array(pf_disposition)
    arr_disposition_1 = np.array(pf_disposition_1)
    dic = dict(arr_disposition_1)
    need_columns = arr_disposition[1]
    col_index = arr_disposition[0]
    wb = Workbook()
    sheet = wb.active

    arr_wight = []
    for i in range(2, len(arr) - 5):
        if pd.isna(arr[i][column_letter_to_number(need_columns[-1]) - 1]):
            sheet.cell(row=i, column=11).value = 2
        else:
            if '单边留' in arr[i][column_letter_to_number(need_columns[-1]) - 1]:
                sheet.cell(row=i, column=11).value = 1
            elif '双边留' in arr[i][column_letter_to_number(need_columns[-1]) - 1]:
                sheet.cell(row=i, column=11).value = 1
                arr_wight.append(i)
            else:
                sheet.cell(row=i, column=11).value = 2
        for j in range(len(need_columns) - 2):
            sheet.cell(row=i, column=j + 1).value = arr[i][column_letter_to_number(need_columns[j]) - 1]
    for k in range(len(arr_wight)):
        sheet.cell(row=int(arr_wight[k]), column=2).value = arr[arr_wight[k]][column_letter_to_number(need_columns[1]) - 1] + arr[arr_wight[k]][
            column_letter_to_number('B') - 1]
    for col_index_num in range(len(col_index)):
        sheet.cell(row=1, column=col_index_num + 1).value = col_index[col_index_num]
        sheet.column_dimensions[get_column_letter(col_index_num + 1)].width = 13.5
    sheet.column_dimensions[get_column_letter(1)].width = 17

    del_num = 0
    for del_row in range(2, len(arr1) - 5):
        if pd.isna(arr1[del_row][column_letter_to_number(arr_disposition[1][5]) - 1]):
            sheet.delete_rows(del_row - del_num)
            del_num += 1

    for i in range(2, len(arr) - 5 - del_num):
        try:
            sheet.cell(row=i, column=len(need_columns) - 1).value = dic[sheet.cell(row=i, column=6).value]
            # print(arr[i][column_letter_to_number(arr_disposition[1][5]) - 1])
        except KeyError:
            sheet.cell(row=i, column=len(need_columns) - 1).value = "未找到"

    wb.save(file.replace('unmerged.xlsx', '') + '压焊参数.xlsx')
    process_excel_file(file.replace('unmerged.xlsx', '') + '压焊参数.xlsx')


def process_excel_file(file_path_3):
    df = pd.read_excel(file_path_3)
    df_transpose = df.transpose()
    df_index = df_transpose.index
    arr = np.array(df)
    wb = Workbook()
    sheet = wb.active
    for j in range(0, len(arr)):
        sheet.column_dimensions[get_column_letter(j * 2 + 1)].width = 14
        for i in range(0, len(df_index)):
            sheet.cell(row=i + 1, column=j * 2 + 1).value = df_index[i]
    for k in range(0, len(arr)):
        sheet.column_dimensions[get_column_letter(k * 2 + 2)].width = 16
        for col_idx in range(0, len(df_index)):
            sheet.cell(row=col_idx + 1, column=k * 2 + 2).value = arr[k][col_idx]
    file_out = file_path_3.replace('.xlsx', '')
    file_out = file_out.replace('.xls', '')
    wb.save(file_out + '_' + '转置.xlsx')
    messagebox.showinfo(message=f"成功!!!")
    root.destroy()
    os.remove(file_path_2)


# 选择
def select_file():
    global file_path_1, options, combobox
    file_path_1 = filedialog.askopenfilename(
        title="选择文件",  # 对话框标题
        filetypes=[("表格", "*.xlsx*")]  # 可选择的文件类型
    )
    if file_path_1:
        path_label.config(text=f"选中的文件：{file_path_1}")
        execl_file = pd.ExcelFile(file_path_1)
        options = execl_file.sheet_names
        combobox = ttk.Combobox(main_app_frame, values=options, width=30)
        combobox.grid(row=4, column=0, columnspan=3, padx=10, pady=2)
    return file_path_1


# 拖拽
def drop_file(event):
    global path_label, file_path_1, options, combobox
    file_path_1 = event.data
    if file_path_1.startswith('{') and file_path_1.endswith('}'):
        file_path_1 = file_path_1[1:-1]
    if file_path_1:
        path_label.config(text=f"选中的文件：{file_path_1}")
        execl_file = pd.ExcelFile(file_path_1)
        options = execl_file.sheet_names
        combobox = ttk.Combobox(main_app_frame, values=options, width=30)
        combobox.grid(row=4, column=0, columnspan=3, padx=10, pady=2)


def run():
    global file_path_1
    global file_path_2
    global sheet, combobox
    sheet = combobox.get()
    if file_path_1 is None:
        messagebox.showerror('错误', message='没有选择要处理的文件')
    elif sheet == '':
        messagebox.showerror('错误', message='没有输入工作表名称')
    else:
        create_next_page(main_app_frame)


def create_next_page(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()
    global file_path_1
    global file_path_2
    global sheet
    file_path_2 = unmerged(file_path_1, sheet)
    os.startfile(file_path_2)
    s_info = ttk.Style()
    s_info.theme_use('vista')

    s_info.configure('TButton', font=('宋体', 12), padding=10)
    s_info.configure('TLabel', font=('宋体', 14), padding=5)
    btu_info = ttk.Button(parent_frame, text='下一步', command=execute)
    text_label = ttk.Label(parent_frame, text="请对打开的unmerged.xlsx点击ctrl+s然后关闭，点击下一步",
                           font=("Microsoft YaHei UI", 14, "bold"),
                           wraplength=400,
                           justify='center',
                           anchor="center")
    text_label.pack(pady=10)
    btu_info.pack(anchor=tk.CENTER)


def create_transpose_tool_page(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def select_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        process_excel_file(file_path)

    def drop_file(event):
        file_path = event.data
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        process_excel_file(file_path)

    title_label = ttk.Label(parent_frame,
                            text="压焊参数转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    title_label.pack(pady=(0, 20))
    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))
    style = ttk.Style()
    style.configure('TButton', font=('微软雅黑', 12), padding=10)
    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()
    btn_1 = ttk.Button(parent_frame, text="返回", command=seven_file, style='TButton')
    btn_1.pack()
    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    footer_label.pack(side='bottom', pady=(20, 0))
    parent_frame.drop_target_register(DND_FILES)
    parent_frame.dnd_bind('<<Drop>>', drop_file)


root = TkinterDnD.Tk()
window_width = 600
window_height = 350
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
root.title("Tool V2.0")

root.geometry(f'{window_width}x{window_height}+{x}+{y}')
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(0, weight=1)
main_app_frame = ttk.Frame(root, padding=20)
main_app_frame.grid(sticky='nsew')
main_app_frame.grid_rowconfigure(0, weight=1)
main_app_frame.grid_rowconfigure(1, weight=1)
main_app_frame.grid_rowconfigure(2, weight=1)
main_app_frame.grid_rowconfigure(3, weight=1)
main_app_frame.grid_rowconfigure(4, weight=1)
main_app_frame.grid_columnconfigure(0, weight=1)
main_app_frame.grid_columnconfigure(1, weight=1)
main_app_frame.grid_columnconfigure(2, weight=1)


def show_home_page():
    for widget in main_app_frame.winfo_children():
        widget.destroy()

    s = ttk.Style()
    s.theme_use('vista')
    s.configure('TButton', font=('宋体', 12), padding=10)
    # 主页标题标签
    home_title_label = ttk.Label(main_app_frame,
                                 text="tool V2.0",
                                 font=('微软雅黑', 18, 'bold'),
                                 foreground='#333333')
    home_title_label.pack(pady=(20, 30))

    # 7号机参数转换
    transpose_btn = ttk.Button(main_app_frame,
                               text="七号机压焊参数",
                               command=lambda: seven_file())
    transpose_btn.pack(pady=10)
    # 拆板
    transpose_btn = ttk.Button(main_app_frame,
                               text="拆板工具",
                               command=lambda: show_spilt_page())
    transpose_btn.pack(pady=10)


def show_spilt_page():
    global entry, path_label, file_path_1, options, combobox

    def spilt_run(file_path):
        sheet_name = combobox.get()
        pf = pd.read_excel(file_path, sheet_name=sheet_name)
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        arr = np.array(pf)
        count = 0
        for i in range(2, len(arr) - 5):
            if (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) % arr[i][
                column_letter_to_number('E') - 1] < arr[i][column_letter_to_number('E') - 1] - arr[0][
                column_letter_to_number('AJ') - 1]:
                num = math.floor(
                    (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) / arr[i][
                        column_letter_to_number('E') - 1] + 1)
            else:
                num = math.ceil(
                    (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) / arr[i][
                        column_letter_to_number('E') - 1] + 1)
            if num >= 36:
                in_column = i + count
                pf = pd.concat([pf.iloc[:in_column + 1], pd.DataFrame([[None] * len(pf.columns)], columns=pf.columns),
                                pf.iloc[in_column + 1:]
                                ], ignore_index=True)
                for j in range(12):
                    pf.iat[in_column + 1, j] = arr[i][j]
                count = count + 1

        arr = np.array(pf)
        ws.insert_rows(len(arr) - 3 - count, amount=count)
        for i in range(count):
            current_row = len(arr) - 3 - count + i
            source_row = current_row - 1
            ws.row_dimensions[current_row].height= 22.5
            for col in range(1, len(arr[0]) + 1):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=current_row, column=col)
                target_cell.font = source_cell.font.copy()
                target_cell.alignment = source_cell.alignment.copy()
                target_cell.fill = source_cell.fill.copy()
                target_cell.border = source_cell.border.copy()
                target_cell.number_format = source_cell.number_format
        steel_num = arr[4][column_letter_to_number('S') - 1]
        model = arr[4][column_letter_to_number('AA') - 1]
        twisted_steel_longth = arr[4][column_letter_to_number('AD') - 1]
        for i in range(3, len(arr) - 4):
            ws.cell(row=i + 1, column=13, value=f'=J{i + 1}')
            ws.cell(row=i + 1, column=14, value=f'=(W{i + 1}-1)*E{i + 1}+C{i + 1}')
            ws.cell(row=i + 1, column=15, value=f'=L{i + 1}')
            ws.cell(row=i + 1, column=16, value=f'=U{i + 1}+Z{i + 1}+AG{i + 1}')
            ws.cell(row=i + 1, column=17,
                    value=f'=IF(OR(AND(MOD(K{i + 1}-C{i + 1},E{i + 1})>$AI$2,MOD(K{i + 1}-C{i + 1},E{i + 1})<(E{i + 1}-$AJ$2)),W{i + 1}=35),"单边留","")')
            ws.cell(row=i + 1, column=18, value=f'=N{i + 1}')
            ws.cell(row=i + 1, column=19, value=steel_num)
            ws.cell(row=i + 1, column=20, value=f'=S{i + 1}*O{i + 1}')
            ws.cell(row=i + 1, column=21, value=f'=B{i + 1}*C{i + 1}*7.85*0.001*R{i + 1}*0.001*T{i + 1}')
            ws.cell(row=i + 1, column=22, value=f'=M{i + 1}-C{i + 1}*2-2')
            ws.cell(row=i + 1, column=23,
                    value=f'=IF(MOD(K{i + 1}-C{i + 1},E{i + 1})<(E{i + 1}-$AJ$2),ROUNDDOWN((K{i + 1}-C{i + 1})/E{i + 1}+1,0),ROUNDUP((K{i + 1}-C{i + 1})/E{i + 1}+1,0))')
            ws.cell(row=i + 1, column=24, value=f'=(N{i + 1}-C{i + 1})/E{i + 1}+1')
            ws.cell(row=i + 1, column=25, value=f'=X{i + 1}*O{i + 1}')
            ws.cell(row=i + 1, column=26, value=f'=B{i + 1}*C{i + 1}*7.85*0.001*V{i + 1}*0.001*Y{i + 1}')
            ws.cell(row=i + 1, column=27, value=model)
            ws.cell(row=i + 1, column=28, value=f'=(M{i + 1}-G{i + 1}*(AE{i + 1}-1))/2')
            ws.cell(row=i + 1, column=29, value=f'=(M{i + 1}-G{i + 1}*(AF{i + 1}-1))/2')
            ws.cell(row=i + 1, column=30, value=twisted_steel_longth)
            ws.cell(row=i + 1, column=31,
                    value=f'=IF(INT(M{i + 1}/G{i + 1}+1)/2=INT(INT(M{i + 1}/G{i + 1}+1)/2),INT(M{i + 1}/G{i + 1})+1,INT(M{i + 1}/G{i + 1}))')
            ws.cell(row=i + 1, column=32, value=f'=IF(AB{i + 1}<=0.2*G{i + 1},AE{i + 1}-2,AE{i + 1})')
            ws.cell(row=i + 1, column=33, value=f'=AA{i + 1}*AA{i + 1}*7.85*AD{i + 1}*0.001*0.001*AF{i + 1}*O{i + 1}')
            ws.cell(row=i + 1, column=34, value=f'=AF{i + 1}*O{i + 1}')
            ws.cell(row=i + 1, column=35, value=f'=K{i + 1}*L{i + 1}')
            ws.cell(row=i + 1, column=36, value=f'=LEFT(H{i + 1},3)')
            for j in range(12):
                ws.cell(i + 1, j + 1).value = arr[i - 1][j]
        font_ins = Font(color='FFFF0000')
        for i in range(3, len(arr) - 4):
            if arr[i - 1][column_letter_to_number('H') - 1] == arr[i][column_letter_to_number('H') - 1]:
                ws.merge_cells(f'H{i + 1}:H{i + 2}')
                ws.cell(i + 1, column_letter_to_number('H')).font = font_ins
                ws.merge_cells(f'I{i + 1}:I{i + 2}')
                ws.merge_cells(f'J{i + 1}:J{i + 2}')
                ws.merge_cells(f'K{i + 1}:K{i + 2}')
                ws.merge_cells(f'L{i + 1}:L{i + 2}')
                ws.cell(i + 2, column_letter_to_number('M')).value = arr[i - 1][column_letter_to_number('M') - 1]
                if (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) % \
                        arr[i - 1][
                            column_letter_to_number('E') - 1] < arr[i - 1][column_letter_to_number('E') - 1] - arr[0][
                    column_letter_to_number('AJ') - 1]:
                    branch_number = (math.floor(
                        (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                        arr[i - 1][column_letter_to_number('E') - 1]) - 34) * arr[i - 1][
                                        column_letter_to_number('E') - 1] + \
                                    arr[i - 1][
                                        column_letter_to_number('C') - 1]
                else:
                    branch_number = (math.ceil(
                        (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                        arr[i - 1][column_letter_to_number('E') - 1]) - 34) * arr[i - 1][
                                        column_letter_to_number('E') - 1] + \
                                    arr[i - 1][
                                        column_letter_to_number('C') - 1]
                if (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) % \
                        arr[i - 1][column_letter_to_number('E') - 1] < arr[i - 1][column_letter_to_number('E') - 1] - \
                        arr[0][
                            column_letter_to_number('AJ') - 1] and (
                        arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) % \
                        arr[i - 1][column_letter_to_number('E') - 1] > arr[0][column_letter_to_number('AI') - 1]:
                    ws.cell(i + 2, column_letter_to_number('N')).value = branch_number + arr[i + 1][
                        column_letter_to_number('E') - 1]
                else:
                    ws.cell(i + 2, column_letter_to_number('N')).value = branch_number
                ws.cell(i + 1, column_letter_to_number('N')).value = 33 * arr[i - 1][4] + arr[i - 1][2]
                ws.cell(i + 2, column_letter_to_number('O')).value = arr[i - 1][column_letter_to_number('O') - 1]
                ws.cell(i + 1, column_letter_to_number('Q')).value = None
                ws.cell(i + 2, column_letter_to_number('Q')).value = '单边留'
            weight_num = 0
            if (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) % \
                    arr[i - 1][
                        column_letter_to_number('E') - 1] < arr[i - 1][column_letter_to_number('E') - 1] - arr[0][
                column_letter_to_number('AJ') - 1]:
                weight_num = (math.floor(
                    (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                    arr[i - 1][column_letter_to_number('E') - 1])) * arr[i - 1][
                                 column_letter_to_number('E') - 1] + \
                             arr[i - 1][
                                 column_letter_to_number('C') - 1]
            else:
                weight_num = (math.ceil(
                    (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                    arr[i - 1][column_letter_to_number('E') - 1])) * arr[i - 1][
                                 column_letter_to_number('E') - 1] + \
                             arr[i - 1][
                                 column_letter_to_number('C') - 1]
            if weight_num < 500:
                ws.cell(i + 1, column_letter_to_number('H')).font = font_ins
                print(weight_num)
        ws.cell(len(arr) - 3, column_letter_to_number('L')).value = f'=SUM(L4:L{len(arr) - 4})'
        ws.cell(len(arr) - 3, column_letter_to_number('O')).value = f'=SUM(O4:O{len(arr) - 4})'
        ws.cell(len(arr) - 1, column_letter_to_number('M')).value = f'=SUM(AI4:AI{len(arr) - 4})/1000*2'
        ws.print_area = f'H1:AH{len(arr) + 1}'
        file_path_out = file_path.replace('.xlsx', '_spilt.xlsx')
        wb.save(f'{file_path_out}')
        messagebox.showinfo(message='拆板完成！')
        root.destroy()

    for widget in main_app_frame.winfo_children():
        widget.destroy()
    s = ttk.Style()
    s.theme_use('vista')
    s.configure('TButton', font=('宋体', 12), padding=10)
    s.configure('TLabel', font=('宋体', 14), padding=5)

    # 使用 ttk 按钮和标签
    btn = ttk.Button(main_app_frame, text="选择文件", command=select_file)
    text_label_1 = ttk.Label(main_app_frame, text="请填写或选择工作表名称(必填)", font=("Microsoft YaHei UI", 14),
                             anchor="center")
    text_label_1.grid(row=3, column=0, columnspan=3, padx=10, pady=2)

    # 输入工作表
    entry = ttk.Entry(main_app_frame, width=20, background="", font=("Microsoft YaHei UI", 14))
    # entry.grid(row=4, column=0, columnspan=3, padx=10, pady=2)

    text_label_2 = ttk.Label(main_app_frame, text="请选择要处理的Excel文件或将文件拖拽到此处",
                             font=("Microsoft YaHei UI", 14, "bold"), anchor="center")
    text_label_2.grid(row=0, column=0, columnspan=3, padx=5, pady=10)

    # 显示路径
    path_label = tk.Label(main_app_frame, text="未选择文件")
    path_label.grid(row=1, column=0, columnspan=3, padx=5, pady=10)

    btn.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', drop_file)

    btn_1 = ttk.Button(main_app_frame, text="确定", command=lambda: spilt_run(file_path_1))
    btn_1.grid(row=5, column=0, columnspan=3, padx=10, pady=10)

    btn_2 = ttk.Button(main_app_frame, text="返回主页", command=show_home_page)
    btn_2.grid(row=5, column=0, columnspan=3, padx=10, pady=10, sticky=tk.W)


def seven_file():
    global entry, path_label
    for widget in main_app_frame.winfo_children():
        widget.destroy()
    # 应用 ttk 主题
    s = ttk.Style()
    s.theme_use('vista')  # 可以尝试 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative' 等

    # 配置按钮和标签样式
    s.configure('TButton', font=('宋体', 12), padding=10)
    s.configure('TLabel', font=('宋体', 14), padding=5)

    # 使用 ttk 按钮和标签
    btn = ttk.Button(main_app_frame, text="选择文件", command=select_file)
    text_label_1 = ttk.Label(main_app_frame, text="请填写或选择工作表名称(必填)", font=("Microsoft YaHei UI", 14),
                             anchor="center")
    text_label_1.grid(row=3, column=0, columnspan=3, padx=10, pady=2)

    # 输入工作表
    entry = ttk.Entry(main_app_frame, width=20, background="", font=("Microsoft YaHei UI", 14))
    # entry.grid(row=4, column=0, columnspan=3, padx=10, pady=2)

    text_label_2 = ttk.Label(main_app_frame, text="请选择要处理的Excel文件或将文件拖拽到此处",
                             font=("Microsoft YaHei UI", 14, "bold"), anchor="center")
    text_label_2.grid(row=0, column=0, columnspan=3, padx=5, pady=10)

    # 显示路径
    path_label = tk.Label(main_app_frame, text="未选择文件")
    path_label.grid(row=1, column=0, columnspan=3, padx=5, pady=10)

    btn.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', drop_file)

    btn_1 = ttk.Button(main_app_frame, text="确定", command=run)
    btn_1.grid(row=5, column=1, columnspan=1, padx=10, pady=10)

    btn_2 = ttk.Button(main_app_frame, text='有压焊参数文件',
                       command=lambda: create_transpose_tool_page(main_app_frame))
    btn_2.grid(row=5, column=2, columnspan=1, padx=20, pady=10, sticky=tk.E)

    btn_3 = ttk.Button(main_app_frame, text='返回首页',
                       command=lambda: show_home_page())
    btn_3.grid(row=5, column=0, columnspan=1, padx=20, pady=10, sticky=tk.W)


show_home_page()
root.mainloop()
