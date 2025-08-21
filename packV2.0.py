import math
import os
import time
import openpyxl
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment
from tkinter import messagebox
from openpyxl.styles import Border, Side


# 拆分最前面的图号字母
def split_string(drawno, split_no):
    arr_drawno = []
    for char in drawno:
        if split_no != 0:
            arr_drawno.append(char)
            split_no -= 1
        else:
            break
    drawno_out = ''.join(arr_drawno)
    return drawno_out


df = pd.read_excel('待分包文件.xlsx')
arr = np.array(df)
steel_height = 60  # 扁钢高度
baseboard_height = 125  # 踢脚板高度
baseboard_thickness = 6  # 踢脚板厚度
floor_height_flag = True  # 层高
split_no = 6  # 包号格式字母范围
qty, weight, insert_count, count1, count2 = 0, 0, 0, 0, 0
density = arr[1][7] / arr[1][6]  # 密度
number = 980 // steel_height
for i in range(len(arr)):
    baseboard_sum = 0
    for j in arr[i][8:-2]:
        baseboard_sum = baseboard_sum + j
        # print(arr[i][1], baseboard_sum)
    if not pd.isna(baseboard_sum) and baseboard_sum != 0:
        count1 += arr[i][5]
        weight += arr[i][7]
        # print(count1,arr[i][1],baseboard_sum)
    else:
        count2 += arr[i][5]
        weight += arr[i][7]
        # print(count2, arr[i][1])
    qty = math.ceil(count1 / 2) * (steel_height + baseboard_height) + count2 * steel_height
    if qty > 980 or weight > 2500:
        # print(arr[i][1], qty, weight)
        baseboard_flag = True


        def moreno(arr, i_no, qty_no, df_no, insert_count_no):
            global count1, count2
            in_column = i_no + insert_count_no
            df_no = pd.concat(
                [df_no.iloc[:in_column], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                 df_no.iloc[in_column:]], ignore_index=True)
            df_no.iloc[in_column] = df_no.iloc[in_column + 1]
            little_length = 0
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                qty_no_last = qty_no - (math.ceil(arr[i_no][5] / 2) * (steel_height + baseboard_height))
                count_last = (980 - qty_no_last) // (steel_height + baseboard_height) * 2

            else:
                # print(qty_no)
                qty_no_last = qty_no - arr[i_no][5] * steel_height
                # print(qty_no_last)
                count_last = (980 - qty_no_last) // steel_height

            df_no.iloc[in_column, 5] = count_last
            # print(count_last)
            df_no.iloc[in_column, 6] = df_no.iloc[in_column, 3] * df_no.iloc[in_column, 4] * df_no.iloc[
                in_column, 5] / 1000000
            df_no.iloc[in_column, 7] = density * df_no.iloc[in_column, 6]
            for have_num in range(8, len(arr[i]) - 2, 2):
                if df_no.iloc[in_column, have_num] != 0:
                    little_length = arr[i_no][have_num] / arr[i_no][5]
                    # print(arr[i_no][1], little_length, arr[i_no][have_num])
                    break
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                for have_num in range(8, len(arr[i]) - 2, 2):
                    # print(have_num)
                    if df_no.iloc[in_column, have_num] != 0:
                        df_no.iloc[in_column, have_num] = little_length * df_no.iloc[in_column, 5]
                        df_no.iloc[in_column, have_num + 1] = df_no.iloc[
                                                                  in_column, have_num] * baseboard_height / 1000 * baseboard_thickness / 1000 * 7.85 * 1.06
            arr[i_no][5] -= df_no.iloc[in_column, 5]
            df_no.iloc[in_column + 1, 5] = arr[i_no][5]
            df_no.iloc[in_column + 1, 6] = df_no.iloc[in_column + 1, 3] * df_no.iloc[in_column + 1, 4] * \
                                           df_no.iloc[
                                               in_column + 1, 5] / 1000000
            df_no.iloc[in_column + 1, 7] = density * df_no.iloc[in_column + 1, 6]
            arr[i][7] = df_no.iloc[in_column + 1, 7]
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                for have_num in range(8, len(arr[i]) - 2, 2):
                    if df_no.iloc[in_column + 1, have_num] != 0:
                        df_no.iloc[in_column + 1, have_num] = little_length * df_no.iloc[in_column + 1, 5]
                        df_no.iloc[in_column + 1, have_num] = df_no.iloc[
                                                                  in_column + 1, have_num] * baseboard_height / 1000 * baseboard_thickness / 1000 * 7.85 * 1.06
            df_no = pd.concat(
                [df_no.iloc[:in_column + 1], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                 df_no.iloc[in_column + 1:]], ignore_index=True)
            insert_count_no += 2
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                qty_no, weight_no = math.ceil(arr[i_no][5] / 2) * (steel_height + baseboard_height), df_no.iloc[
                    in_column + 1, 7]
                count1, count2 = arr[i][5], 0
                # print(qty_no, arr[i_no][1], arr[i_no][5])
            else:
                qty_no, weight_no = arr[i_no][5] * steel_height, df_no.iloc[
                    in_column + 1, 7]
                count1, count2 = 0, arr[i][5]
                # print(qty_no, arr[i_no][1], arr[i_no][5])
            return qty_no, weight_no, insert_count_no, df_no


        def moreweight(arr, i_no, weight_no, df_no, insert_count_no):
            global count1, count2

            in_column = i_no + insert_count_no
            df_no = pd.concat(
                [df_no.iloc[:in_column], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                 df_no.iloc[in_column:]], ignore_index=True)
            df_no.iloc[in_column] = df_no.iloc[in_column + 1]
            little_length = 0
            print(arr[i_no][7], weight_no)
            count_last = math.floor(
                (arr[i_no][7] - (weight_no - 2500)) / density / arr[i_no][3] / arr[i_no][4] * 1000000)
            print(count_last)
            df_no.iloc[in_column, 5] = count_last
            df_no.iloc[in_column, 6] = df_no.iloc[in_column, 3] * df_no.iloc[in_column, 4] * df_no.iloc[
                in_column, 5] / 1000000
            df_no.iloc[in_column, 7] = density * df_no.iloc[in_column, 6]
            for have_num in range(8, len(arr[i]) - 2, 2):
                if df_no.iloc[in_column, have_num] != 0:
                    little_length = arr[i_no][have_num] / arr[i_no][5]
                    break
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                for have_num in range(8, len(arr[i]) - 2, 2):
                    if df_no.iloc[in_column, have_num] != 0:
                        df_no.iloc[in_column, have_num] = little_length * df_no.iloc[in_column, 5]
                        df_no.iloc[in_column, have_num + 1] = df_no.iloc[
                                                                  in_column, have_num] * baseboard_height / 1000 * baseboard_thickness / 1000 * 7.85 * 1.06
            arr[i_no][5] -= df_no.iloc[in_column, 5]
            arr[i_no][7] = arr[i_no][4] * arr[i_no][5] * arr[i_no][3] / 1000000 * density
            df_no.iloc[in_column + 1, 5] = arr[i_no][5]
            df_no.iloc[in_column + 1, 6] = df_no.iloc[in_column + 1, 3] * df_no.iloc[in_column + 1, 4] * \
                                           df_no.iloc[
                                               in_column + 1, 5] / 1000000
            df_no.iloc[in_column + 1, 7] = density * df_no.iloc[in_column + 1, 6]
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                for have_num in range(8, len(arr[i]) - 2, 2):
                    if df_no.iloc[in_column + 1, have_num] != 0:
                        df_no.iloc[in_column + 1, have_num] = little_length * df_no.iloc[in_column + 1, 5]
                        df_no.iloc[in_column + 1, have_num] = df_no.iloc[
                                                                  in_column + 1, have_num] * baseboard_height / 1000 * baseboard_thickness / 1000 * 7.85 * 1.06
            df_no = pd.concat(
                [df_no.iloc[:in_column + 1], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                 df_no.iloc[in_column + 1:]], ignore_index=True)
            insert_count_no += 2
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                qty_no, weight_no = math.ceil(arr[i_no][5] / 2) * (steel_height + baseboard_height), df_no.iloc[
                    in_column + 1, 7]
                count1, count2 = arr[i][5], 0
            else:
                print(1,df_no.iloc[in_column + 2, 7])
                qty_no, weight_no = arr[i_no][5] * steel_height, df_no.iloc[in_column + 2, 7]
                count1, count2 = 0, arr[i][5]
            return qty_no, weight_no, insert_count_no, df_no


        if not pd.isna(baseboard_sum) and baseboard_sum != 0:
            while math.ceil(arr[i][5] / 2) * (steel_height + baseboard_height) > 980 or arr[i][7] > 2500:
                if arr[i][7] > 2500:
                    qty, weight, insert_count, df = moreweight(arr, i, weight, df, insert_count)
                else:
                    qty, weight, insert_count, df = moreno(arr, i, qty, df, insert_count)
                baseboard_flag = False

        else:
            while arr[i][5] * steel_height > 980 or arr[i][7] > 2500:
                if arr[i][7] > 2500:
                    qty, weight, insert_count, df = moreweight(arr, i, weight, df, insert_count)
                else:
                    qty, weight, insert_count, df = moreno(arr, i, qty, df, insert_count)
                baseboard_flag = False
        if baseboard_flag:
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                count1, count2, weight = arr[i][5], 0, arr[i][7]
            else:
                count1, count2, weight = 0, arr[i][5], arr[i][7]


    elif i > 1:
        if split_string(arr[i][1], split_no) != split_string(arr[i - 1][1], split_no):
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                count1, count2, weight = arr[i][5], 0, arr[i][7]
            else:
                count1, count2, weight = 0, arr[i][5], arr[i][7]
        elif floor_height_flag == True and arr[i][-1] != arr[i - 1][-1]:
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                count1, count2, weight = arr[i][5], 0, arr[i][7]
            else:
                count1, count2, weight = 0, arr[i][5], arr[i][7]
arr = np.array(df)
insert_count = 0
for i in range(1, len(arr)):
    if not (pd.isna(arr[i][1]) or pd.isna(arr[i - 1][1])):
        if split_string(arr[i][1], split_no) != split_string(arr[i - 1][1], split_no):
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                count1, count2, weight = arr[i][5], 0, arr[i][7]
            else:
                count1, count2, weight = 0, arr[i][5], arr[i][7]
        elif floor_height_flag == True and arr[i][-1] != arr[i - 1][-1]:
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            if not pd.isna(baseboard_sum) and baseboard_sum != 0:
                count1, count2, weight = arr[i][5], 0, arr[i][7]
            else:
                count1, count2, weight = 0, arr[i][5], arr[i][7]

arr = np.array(df)
qty, pre_empty, set_beyond, count1, count2 = 0, 0, 180, 0, 0
for i in range(len(arr)):
    if arr[i][5] == 0:
        df.drop(index=i, inplace=True)
    if not arr[i][1] is None:
        for j in arr[i][8:-2]:
            baseboard_sum = baseboard_sum + j
        if not pd.isna(baseboard_sum) and baseboard_sum != 0:
            count1 += arr[i][5]
        else:
            count2 += arr[i][5]
        # print(count1, count2)
        qty = math.ceil(count1 / 2) * (steel_height + baseboard_height) + count2 * steel_height
        # print(qty, arr[i][1])
    else:
        # print(arr[i - 1][1], qty)
        # if qty < set_beyond:
        #     if pre_empty < 5:
        #
        #         if arr[i - 1][-1] != arr[i + 1][-1] and floor_height_flag == True:
        #             pass
        #         else:
        #             df.drop(index=i, inplace=True)
        #     else:
        #         if arr[pre_empty - 1][-1] != arr[pre_empty + 1][-1] and floor_height_flag == True:
        #             pass
        #         else:
        #             df.drop(index=pre_empty, inplace=True)
        count1, count2 = 0, 0
        pre_empty = i
pack_no = ''
pack_no_num = 1
pack_no_arr = ['01', '02', '03', '04', '05', '06', '07', '08', '09']
pre_loc = 1
df.loc[len(arr) + 3] = None
arr = np.array(df)
color_arr = []
for i in range(len(arr)):
    if pack_no_num < 10:
        pack_no = pack_no_arr[pack_no_num - 1]
        df.iloc[i, 0] = f'P{pack_no}'
    else:
        df.iloc[i, 0] = f'P{pack_no_num}'
    if arr[i][1] is None or pd.isna(arr[i][1]):
        color_arr.append(i)
        df.iloc[i, 0] = f'小计'
        pack_no_num += 1
        df.iloc[i, 5] = f'=sum(F{pre_loc + 1}:F{i + 1})'
        df.iloc[i, 7] = f'=sum(H{pre_loc + 1}:H{i + 1})'
        pre_loc = i + 2
end_i = color_arr[-1] + 2
color_arr.append(end_i)

df.to_excel("out.xlsx", index=False)

wb = load_workbook("out.xlsx")
ws = wb.active
fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # 黄色

for row_index in color_arr:
    for col_index in range(1, ws.max_column + 1):
        ws.cell(row=row_index + 2, column=col_index).fill = fill
ws.row_dimensions[color_arr[-1] + 1].height = 10
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for cell_row in ws[1]:
    cell_row.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
for i in range(2, len(arr) + 1):
    for cell_row in ws[i]:
        cell_row.alignment = Alignment(horizontal='center', vertical='center')
        cell_row.border = border
ws.cell(row=color_arr[-1] + 2, column=1).value = f'合计'
ws.cell(row=color_arr[-1] + 2, column=6).value = f'=sum(f2:f{color_arr[-1]})/2'
for row_num in range(7, len(ws[1]) - 1):
    for col_index in range(2, len(arr) + 1):
        cell_num = ws.cell(row=col_index, column=row_num)
        cell_num.number_format = '0.00'
# 保存修改后的工作簿
wb.save("out.xlsx")

os.startfile("out.xlsx")
