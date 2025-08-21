import math
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter


# """将表格列字母（如'A'、'AB'）转换为数字（如1、28）"""
def column_letter_to_number(letter):
    number = 0
    for char in letter.upper():
        number = number * 26 + (ord(char) - ord('A') + 1)
    return number


file_path = '7.xlsx'
sheet_name = '8成品二算料单标准（修改版）'
pf = pd.read_excel(file_path, sheet_name=sheet_name)
wb = openpyxl.load_workbook(file_path)
ws = wb[sheet_name]
arr = np.array(pf)
count = 0
for i in range(2, len(arr) - 5):
    num = 0
    if (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) % arr[i][
        column_letter_to_number('E') - 1] < arr[i][column_letter_to_number('E') - 1] - arr[0][
        column_letter_to_number('AJ') - 1]:
        num = math.floor(
            (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) / arr[i][
                column_letter_to_number('E') - 1] + 1)
    else:
        num = math.ceil(
            (arr[i][column_letter_to_number('K') - 1] - arr[i][column_letter_to_number('C') - 1]) / arr[i][
                column_letter_to_number('E') - 1] + 1)
    if num >= 36:
        print(num)
        in_column = i + count
        pf = pd.concat([pf.iloc[:in_column + 1], pd.DataFrame([[None] * len(pf.columns)], columns=pf.columns),
                        pf.iloc[in_column + 1:]
                        ], ignore_index=True)
        for j in range(12):
            pf.iat[in_column + 1, j] = arr[i][j]
        count = count + 1
        print(count)

arr = np.array(pf)
print()
ws.insert_rows(len(arr) - 3 - count, amount=count)
steel_num = arr[4][column_letter_to_number('S') - 1]
model = arr[4][column_letter_to_number('AA') - 1]
twisted_steel_longth = arr[4][column_letter_to_number('AD') - 1]
for i in range(3, len(arr) - 4):
    ws.cell(row=i + 1, column=13, value=f'=J{i + 1}')
    ws.cell(row=i + 1, column=14, value=f'=(W{i + 1}-1)*E{i + 1}+C{i + 1}')
    ws.cell(row=i + 1, column=15, value=f'=L{i + 1}')
    ws.cell(row=i + 1, column=16, value=f'=U{i + 1}+Z{i + 1}+AG{i + 1}')
    ws.cell(row=i + 1, column=17,
            value=f'=IF(OR(AND(MOD(K{i + 1}-C{i + 1},E{i + 1})>$AI$2,MOD(K{i + 1}-C{i + 1},E{i + 1})<(E{i + 1}-$AJ$2)),W{i + 1}=35),"单边留","")')
    ws.cell(row=i + 1, column=18, value=f'=N{i + 1}')
    ws.cell(row=i + 1, column=19, value=steel_num)
    ws.cell(row=i + 1, column=20, value=f'=S{i + 1}*O{i + 1}')
    ws.cell(row=i + 1, column=21, value=f'=B{i + 1}*C{i + 1}*7.85*0.001*R{i + 1}*0.001*T{i + 1}')
    ws.cell(row=i + 1, column=22, value=f'=M{i + 1}-C{i + 1}*2-2')
    ws.cell(row=i + 1, column=23,
            value=f'=IF(MOD(K{i + 1}-C{i + 1},E{i + 1})<(E{i + 1}-$AJ$2),ROUNDDOWN((K{i + 1}-C{i + 1})/E{i + 1}+1,0),ROUNDUP((K{i + 1}-C{i + 1})/E{i + 1}+1,0))')
    ws.cell(row=i + 1, column=24, value=f'=(N{i + 1}-C{i + 1})/E{i + 1}+1')
    ws.cell(row=i + 1, column=25, value=f'=X{i + 1}*O{i + 1}')
    ws.cell(row=i + 1, column=26, value=f'=B{i + 1}*C{i + 1}*7.85*0.001*V{i + 1}*0.001*Y{i + 1}')
    ws.cell(row=i + 1, column=27, value=model)
    ws.cell(row=i + 1, column=28, value=f'=(M{i + 1}-G{i + 1}*(AE{i + 1}-1))/2')
    ws.cell(row=i + 1, column=29, value=f'=(M{i + 1}-G{i + 1}*(AF{i + 1}-1))/2')
    ws.cell(row=i + 1, column=30, value=twisted_steel_longth)
    ws.cell(row=i + 1, column=31,
            value=f'=IF(INT(M{i + 1}/G{i + 1}+1)/2=INT(INT(M{i + 1}/G{i + 1}+1)/2),INT(M{i + 1}/G{i + 1})+1,INT(M{i + 1}/G{i + 1}))')
    ws.cell(row=i + 1, column=32, value=f'=IF(AB{i + 1}<=0.2*G{i + 1},AE{i + 1}-2,AE{i + 1})')
    ws.cell(row=i + 1, column=33, value=f'=AA{i + 1}*AA{i + 1}*7.85*AD{i + 1}*0.001*0.001*AF{i + 1}*O{i + 1}')
    ws.cell(row=i + 1, column=34, value=f'=AF{i + 1}*O{i + 1}')
    for j in range(12):
        ws.cell(i + 1, j + 1).value = arr[i - 1][j]
for i in range(3, len(arr) - 4):
    if arr[i - 1][column_letter_to_number('H') - 1] == arr[i][column_letter_to_number('H') - 1]:
        ws.merge_cells(f'H{i + 1}:H{i + 2}')
        ws.merge_cells(f'I{i + 1}:I{i + 2}')
        ws.merge_cells(f'J{i + 1}:J{i + 2}')
        ws.merge_cells(f'K{i + 1}:K{i + 2}')
        ws.merge_cells(f'L{i + 1}:L{i + 2}')
        ws.cell(i + 2, column_letter_to_number('M')).value = arr[i - 1][column_letter_to_number('M') - 1]
        if (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) % arr[i - 1][
            column_letter_to_number('E') - 1] < arr[i - 1][column_letter_to_number('E') - 1] - arr[0][
            column_letter_to_number('AJ') - 1]:
            branch_number = (math.floor(
                (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                arr[i - 1][column_letter_to_number('E')]) - 34) * arr[i - 1][column_letter_to_number('E')] + arr[i - 1][
                                column_letter_to_number('C')]
        else:
            branch_number = (math.ceil(
                (arr[i - 1][column_letter_to_number('K') - 1] - arr[i - 1][column_letter_to_number('C') - 1]) /
                arr[i - 1][column_letter_to_number('E')]) - 34) * arr[i - 1][column_letter_to_number('E')] + arr[i - 1][
                                column_letter_to_number('C')]
        ws.cell(i + 2, column_letter_to_number('N')).value = branch_number
        ws.cell(i + 1, column_letter_to_number('N')).value = 33 * arr[i - 1][4] + arr[i - 1][2]
        ws.cell(i + 2, column_letter_to_number('O')).value = arr[i - 1][column_letter_to_number('O') - 1]
        ws.cell(i + 1, column_letter_to_number('Q')).value = None
        ws.cell(i + 2, column_letter_to_number('Q')).value = '单边留'
ws.print_area = f'H1:AH{len(arr) + 1}'

wb.save('out.xlsx')
