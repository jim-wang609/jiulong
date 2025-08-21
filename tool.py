import os
# import time
import tkinter as tk
from tkinter import ttk  # 导入ttk模块
from tkinter import filedialog  # 导入文件选择模块
from tkinter import messagebox

import openpyxl
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tkinterdnd2 import TkinterDnD, DND_FILES

# 全局变量
file_path_1 = None
sheet = None
entry = None
path_label = None


# """将表格列字母（如'A'、'AB'）转换为数字（如1、28）"""
def column_letter_to_number(letter):
    number = 0
    for char in letter.upper():
        number = number * 26 + (ord(char) - ord('A') + 1)
    return number


# 输入工作表
def get_input():
    global entry, sheet
    sheet = entry.get()


# 拆分单元格
def unmerged(file_path, sheet):
    if file_path:
        try:
            book = load_workbook(file_path)
            ws = book[sheet]
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell_value = ws.cell(row=min_row, column=min_col).value
                ws.unmerge_cells(str(merged_range))
                for row_idx in range(min_row, max_row + 1):
                    for col_idx in range(min_col, max_col + 1):
                        ws.cell(row=row_idx, column=col_idx).value = top_left_cell_value
            output_file_path = file_path.replace('.xlsx', '') + '_' + "unmerged.xlsx"
            book.save(output_file_path)
            return output_file_path
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")


# 参数主函数
def execute():
    global file_path_2, sheet
    parameter(file_path_2, sheet, file_path_1)


def parameter(file, sheet, file_2):
    pf = pd.read_excel(file, sheet_name=sheet)
    pf1 = pd.read_excel(file_2, sheet_name=sheet)
    pf_disposition = pd.read_excel('disposition.xlsx', sheet_name="Sheet1")
    pf_disposition_1 = pd.read_excel('disposition.xlsx', sheet_name="Sheet2")
    arr = np.array(pf)
    arr1 = np.array(pf1)
    arr_disposition = np.array(pf_disposition)
    arr_disposition_1 = np.array(pf_disposition_1)
    dic = dict(arr_disposition_1)
    need_columns = arr_disposition[1]
    col_index = arr_disposition[0]
    wb = Workbook()
    sheet = wb.active

    for i in range(2, len(arr) - 5):
        if pd.isna(arr[i][column_letter_to_number(need_columns[-1]) - 1]):
            sheet.cell(row=i, column=11).value = 2
        else:
            if '单边留' in arr[i][column_letter_to_number(need_columns[-1]) - 1]:
                sheet.cell(row=i, column=11).value = 1
            else:
                sheet.cell(row=i, column=11).value = 2
        for j in range(len(need_columns) - 2):
            sheet.cell(row=i, column=j + 1).value = arr[i][column_letter_to_number(need_columns[j]) - 1]

    for col_index_num in range(len(col_index)):
        sheet.cell(row=1, column=col_index_num + 1).value = col_index[col_index_num]
        sheet.column_dimensions[get_column_letter(col_index_num + 1)].width = 13.5
    sheet.column_dimensions[get_column_letter(1)].width = 17

    del_num = 0
    for del_row in range(2, len(arr1) - 5):
        if pd.isna(arr1[del_row][column_letter_to_number(arr_disposition[1][5]) - 1]):
            sheet.delete_rows(del_row - del_num)
            del_num += 1

    for i in range(2, len(arr) - 5 - del_num):
        try:
            sheet.cell(row=i, column=len(need_columns) - 1).value = dic[sheet.cell(row=i, column=6).value]
            print(arr[i][column_letter_to_number(arr_disposition[1][5]) - 1])
        except KeyError:
            sheet.cell(row=i, column=len(need_columns) - 1).value = "未找到"

    wb.save(file.replace('unmerged.xlsx', '') + '压焊参数.xlsx')
    process_excel_file(file.replace('unmerged.xlsx', '') + '压焊参数.xlsx')


def process_excel_file(file_path_3):
    df = pd.read_excel(file_path_3)
    df_transpose = df.transpose()
    df_index = df_transpose.index
    arr = np.array(df)
    wb = Workbook()
    sheet = wb.active
    for j in range(0, len(arr)):
        sheet.column_dimensions[get_column_letter(j * 2 + 1)].width = 14
        for i in range(0, len(df_index)):
            sheet.cell(row=i + 1, column=j * 2 + 1).value = df_index[i]
    for k in range(0, len(arr)):
        sheet.column_dimensions[get_column_letter(k * 2 + 2)].width = 16
        for col_idx in range(0, len(df_index)):
            sheet.cell(row=col_idx + 1, column=k * 2 + 2).value = arr[k][col_idx]
    file_out = file_path_3.replace('.xlsx', '')
    file_out = file_out.replace('.xls', '')
    wb.save(file_out + '_' + '转置.xlsx')
    messagebox.showinfo(message=f"成功!!!")
    root.destroy()
    os.remove(file_path_2)


# 选择
def select_file():
    global file_path_1, path_label
    file_path_1 = filedialog.askopenfilename(
        title="选择文件",  # 对话框标题
        filetypes=[("表格", "*.xlsx*")]  # 可选择的文件类型
    )
    if file_path_1:
        path_label.config(text=f"选中的文件：{file_path_1}")
    return file_path_1


# 拖拽
def drop_file(event):
    global path_label, file_path_1
    file_path_1 = event.data
    if file_path_1.startswith('{') and file_path_1.endswith('}'):
        file_path_1 = file_path_1[1:-1]
    if file_path_1:
        path_label.config(text=f"选中的文件：{file_path_1}")


def run():
    global file_path_1
    global file_path_2
    global sheet
    get_input()
    if file_path_1 is None:
        messagebox.showerror('错误', message='没有选择要处理的文件')
    elif sheet == '':
        messagebox.showerror('错误', message='没有输入工作表名称')
    else:
        create_next_page(main_app_frame)


def create_next_page(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()
    global file_path_1
    global file_path_2
    global sheet
    file_path_2 = unmerged(file_path_1, sheet)
    os.startfile(file_path_2)
    s_info = ttk.Style()
    s_info.theme_use('vista')

    s_info.configure('TButton', font=('宋体', 12), padding=10)
    s_info.configure('TLabel', font=('宋体', 14), padding=5)
    btu_info = ttk.Button(parent_frame, text='下一步', command=execute)
    text_label = ttk.Label(parent_frame, text="请对打开的unmerged.xlsx点击ctrl+s然后关闭，点击下一步",
                           font=("Microsoft YaHei UI", 14, "bold"),
                           wraplength=400,
                           justify='center',
                           anchor="center")
    text_label.pack(pady=10)
    btu_info.pack(anchor=tk.CENTER)


def create_transpose_tool_page(parent_frame):
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def select_file():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        process_excel_file(file_path)

    def drop_file(event):
        file_path = event.data
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        process_excel_file(file_path)

    title_label = ttk.Label(parent_frame,
                            text="压焊参数转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    title_label.pack(pady=(0, 20))
    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))
    style = ttk.Style()
    style.configure('TButton', font=('微软雅黑', 12), padding=10)
    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()
    btn_1 = ttk.Button(parent_frame, text="返回首页", command=show_home_page, style='TButton')
    btn_1.pack()
    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    footer_label.pack(side='bottom', pady=(20, 0))
    parent_frame.drop_target_register(DND_FILES)
    parent_frame.dnd_bind('<<Drop>>', drop_file)


root = TkinterDnD.Tk()
window_width = 600
window_height = 350
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
root.title("压焊参数整理")

root.geometry(f'{window_width}x{window_height}+{x}+{y}')
root.grid_columnconfigure(0, weight=1)
root.grid_rowconfigure(0, weight=1)
main_app_frame = ttk.Frame(root, padding=20)
main_app_frame.grid(sticky='nsew')
main_app_frame.grid_rowconfigure(0, weight=1)
main_app_frame.grid_rowconfigure(1, weight=1)
main_app_frame.grid_rowconfigure(2, weight=1)
main_app_frame.grid_rowconfigure(3, weight=1)
main_app_frame.grid_rowconfigure(4, weight=1)
main_app_frame.grid_columnconfigure(0, weight=1)
main_app_frame.grid_columnconfigure(1, weight=1)


def show_home_page():
    global entry, path_label
    for widget in main_app_frame.winfo_children():
        widget.destroy()
    # 应用 ttk 主题
    s = ttk.Style()
    s.theme_use('vista')  # 可以尝试 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative' 等

    # 配置按钮和标签样式
    s.configure('TButton', font=('宋体', 12), padding=10)
    s.configure('TLabel', font=('宋体', 14), padding=5)

    # 使用 ttk 按钮和标签
    btn = ttk.Button(main_app_frame, text="选择文件", command=select_file)
    text_label_1 = ttk.Label(main_app_frame, text="请填写工作表名称(必填)", font=("Microsoft YaHei UI", 14),
                             anchor="center")
    text_label_1.grid(row=3, column=0, columnspan=2, padx=10, pady=2)

    # 输入工作表
    entry = ttk.Entry(main_app_frame, width=20, background="", font=("Microsoft YaHei UI", 14))
    entry.grid(row=4, column=0, columnspan=2, padx=10, pady=2)

    text_label_2 = ttk.Label(main_app_frame, text="请选择要处理的Excel文件或将文件拖拽到此处",
                             font=("Microsoft YaHei UI", 14, "bold"), anchor="center")
    text_label_2.grid(row=0, column=0, columnspan=2, padx=5, pady=10)

    # 显示路径
    path_label = tk.Label(main_app_frame, text="未选择文件")
    path_label.grid(row=1, column=0, columnspan=2, padx=5, pady=10)

    btn.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', drop_file)

    btn_1 = ttk.Button(main_app_frame, text="确定", command=run)
    btn_1.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

    btn_2 = ttk.Button(main_app_frame, text='有压焊参数文件',
                       command=lambda: create_transpose_tool_page(main_app_frame))
    btn_2.grid(row=5, column=1, columnspan=2, padx=20, pady=10, sticky=tk.E)


show_home_page()
root.mainloop()
