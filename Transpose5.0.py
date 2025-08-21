import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import messagebox
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
import os
from tkinterdnd2 import TkinterDnD, DND_FILES

# import time


def process_excel_file(file_path):
    if file_path:
        try:
            df = pd.read_excel(file_path)

            df_transpose = df.transpose()
            df_index = df_transpose.index

            arr = np.array(df)

            wb = Workbook()
            sheet = wb.active
            for j in range(0, len(arr)):
                sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
                for i in range(0, len(df_index)):
                    sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
            for k in range(0, len(arr)):
                sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
                for col_idx in range(0, len(df_index)):
                    sheet.cell(row=col_idx+1, column=k*2+2).value = arr[k][col_idx]

            file_out = file_path.replace('.xlsx', '')
            file_out = file_out.replace('.xls', '')
            wb.save(file_out+'_'+'转置.xlsx')
            messagebox.showinfo(message=f"转置成功！！\n生成文件路径：{file_out}_转置.xlsx")
            root.destroy()
        except Exception as e:
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")


def select_file():
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        initialdir=os.path.expanduser('~'),
        filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
    process_excel_file(file_path)


def drop_file(event):
    file_path = event.data
    if file_path.startswith('{') and file_path.endswith('}'):
        file_path = file_path[1:-1]
    process_excel_file(file_path)


root = TkinterDnD.Tk()
root.title("Excel转置工具")
root.geometry('500x300+500+300')
root.resizable(False, False)
root.configure(bg='#f0f0f0')

main_frame = ttk.Frame(root, padding=20)
main_frame.pack(fill=tk.BOTH, expand=True)

title_label = ttk.Label(main_frame,
                        text="Excel转置工具",
                        font=('微软雅黑', 16, 'bold'),
                        foreground='#333333')
title_label.pack(pady=(0, 20))

text_label = ttk.Label(main_frame,
                       text="请选择要转置的Excel文件或将文件拖拽到此处",
                       font=('微软雅黑', 12))
text_label.pack(pady=(0, 30))

style = ttk.Style()
style.configure('TButton', font=('微软雅黑', 12), padding=10)

btn = ttk.Button(main_frame,
                 text="选择文件",
                 command=select_file,
                 style='TButton')
btn.pack()

footer_label = ttk.Label(main_frame,
                         text="支持.xls和.xlsx格式",
                         font=('微软雅黑', 9),
                         foreground='#666666')
footer_label.pack(side=tk.BOTTOM, pady=(20, 0))

root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', drop_file)

root.mainloop()
