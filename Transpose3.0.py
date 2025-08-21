import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import messagebox
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
import os
# import time


def select_file():
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        initialdir=os.path.expanduser('~'),
        filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")]
    )
    if file_path:  # 如果选择了文件
        file_in = file_path
        df = pd.read_excel(file_in)

        df_transpose = df.transpose()
        df_index = df_transpose.index

        arr = np.array(df)
        # arr_transpose = arr.transpose()
        #print(arr)
        # print(arr_transpose)
        # print(df_transpose.columns)
        # print(df_values)
        # print(df_index)
        # print(df.values)
        wb = Workbook()
        sheet = wb.active
        for j in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
            for i in range(0, len(df_index)):
                sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
        for k in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
            for l in range(0, len(df_index)):
                sheet.cell(row=l+1, column=k*2+2).value = arr[k][l]

        file_out = file_in.replace('.xlsx', '')
        file_out = file_out.replace('.xls', '')
        wb.save(file_out+'_'+'转置.xlsx')
        messagebox.showinfo(message="转置成功！！")
        root.destroy()

# 创建窗口和按钮
root = tk.Tk()
root.title("Excel转置工具")
#root.iconbitmap("图标.ico")
root.geometry('500x300+500+300')
root.resizable(False, False)
root.configure(bg='#f0f0f0')

# 创建主框架
main_frame = ttk.Frame(root, padding=20)
main_frame.pack(fill=tk.BOTH, expand=True)

# 添加标题
title_label = ttk.Label(main_frame, 
                      text="Excel转置工具", 
                      font=('微软雅黑', 16, 'bold'),
                      foreground='#333333')
title_label.pack(pady=(0, 20))

# 添加说明文本
text_label = ttk.Label(main_frame, 
                     text="请选择要转置的Excel文件", 
                     font=('微软雅黑', 12))
text_label.pack(pady=(0, 30))

# 添加选择文件按钮
style = ttk.Style()
style.configure('TButton', font=('微软雅黑', 12), padding=10)

btn = ttk.Button(main_frame, 
               text="选择文件", 
               command=select_file,
               style='TButton')
btn.pack()

# 添加底部信息
footer_label = ttk.Label(main_frame, 
                       text="支持.xls和.xlsx格式", 
                       font=('微软雅黑', 9),
                       foreground='#666666')
footer_label.pack(side=tk.BOTTOM, pady=(20, 0))


def select_file():
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        initialdir=os.path.expanduser('~'),
        filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
    if file_path:  # 如果选择了文件
        file_in = file_path
        df = pd.read_excel(file_in)

        df_transpose = df.transpose()
        df_index = df_transpose.index

        arr = np.array(df)
        # arr_transpose = arr.transpose()
        #print(arr)
        # print(arr_transpose)
        # print(df_transpose.columns)
        # print(df_values)
        # print(df_index)
        # print(df.values)
        wb = Workbook()
        sheet = wb.active
        for j in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
            for i in range(0, len(df_index)):
                sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
        for k in range(0, len(arr)):
            sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
            for l in range(0, len(df_index)):
                sheet.cell(row=l+1, column=k*2+2).value = arr[k][l]

        file_out = file_in.replace('.xlsx', '')
        file_out = file_out.replace('.xls', '')
        wb.save(file_out+'_'+'转置.xlsx')
        messagebox.showinfo(message="转置成功！！")
        root.destroy()

root.mainloop()
