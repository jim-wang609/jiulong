import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import qrcode
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage

row_max = 95
page_max = 10
page_column_max = 3
page_zoom = 75
filepath = 'QR.xlsx'
sheet_name = '7成品一原板套料'
df = pd.read_excel(filepath, sheet_name=sheet_name)
wb = openpyxl.load_workbook(filepath)
ws = wb.create_sheet(title='生成二维码')
# wb = Workbook()
# ws = wb.active
ws.cell(row=1, column=1).value = df.iloc[0, 2]
ws.cell(row=2, column=1).value = df.iloc[0, 10]
pre_num, num_count = df.iloc[7, 0], 0
new_i, new_j, i_count, j_count = 4, 1, 0, 0


def generate_qrcode_image(data, size):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    transparent_bg = True
    if transparent_bg:
        datas = img.getdata()
        new_data = []
        for item in datas:
            if item[0] == 255 and item[1] == 255 and item[2] == 255:
                new_data.append((255, 255, 255, 0))
            else:
                new_data.append(item)
        img.putdata(new_data)

    img = img.resize((size, size), PILImage.Resampling.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return Image(buffer)


for i in range(7, row_max - 1):
    print(df.iloc[i, 4])
    if pd.isna(df.iloc[i, 4]) or df.iloc[i, 4] == " ":
        pre_num = df.iloc[i + 1, 0]
    elif df.iloc[i, 4] == "合并板":
        pass
    else:
        ws.cell(row=new_i + page_max * i_count, column=new_j + 3 * j_count).value = pre_num
        ws.cell(row=new_i + page_max * i_count, column=new_j + 3 * j_count + 1).value = df.iloc[i, 4]
        qr_img = generate_qrcode_image(df.iloc[i, 4], 120)
        target_cell = ws.cell(row=new_i + page_max * i_count, column=new_j + 3 * j_count + 2)
        qr_img.anchor = f"{target_cell.column_letter}{target_cell.row}"
        ws.add_image(qr_img)
        new_i += 1
        if new_i == page_max + 4:
            new_i = 4
            j_count += 1
        if j_count == 3:
            j_count = 0
            i_count += 1
for i in range(3):
    ws.column_dimensions[get_column_letter(1 + 3 * i)].width = 11.3
    ws.column_dimensions[get_column_letter(2 + 3 * i)].width = 19.89
    ws.column_dimensions[get_column_letter(3 + 3 * i)].width = 17
    ws.cell(row=3, column=1 + 3 * i).value = df.iloc[6, 0]
    ws.cell(row=3, column=2 + 3 * i).value = df.iloc[6, 4]
    ws.cell(row=3, column=3 + 3 * i).value = ""
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=4):
    for cell_i in row:
        cell_i.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        cell_i.border = border
        cell_i.font = Font(size=20)
for row in ws.iter_rows(min_row=1, max_row=3):
    for cell_i in row:
        cell_i.alignment = Alignment(horizontal='center', vertical='center')
        cell_i.border = border
        cell_i.font = Font(size=18)
for row in range(1, 4):
    ws.row_dimensions[row].height = 22.5
for row in range(4, ws.max_row + 1):
    ws.row_dimensions[row].height = 89
ws.page_margins.left = 0.03
ws.page_margins.right = 0
ws.page_margins.top = 0.2
ws.page_margins.bottom = 0
ws.print_title_rows = '1:3'
ws.print_area = f'A1:i{ws.max_row}'
ws.page_setup.scale = 75
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.merge_cells('A1:I1')
ws.merge_cells('A2:I2')
wb.save(filepath)
os.startfile(filepath)
