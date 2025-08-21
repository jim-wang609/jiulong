import os
import time
import openpyxl
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 拆分最前面的图号字母
def split_string(drawno, split_no):
    arr_drawno = []
    for char in drawno:
        if split_no != 0:
            arr_drawno.append(char)
            split_no -= 1
        else:
            break
    drawno_out = ''.join(arr_drawno)
    return drawno_out


df = pd.read_excel('待分包文件.xlsx')
arr = np.array(df)
steel_height = 25  # 扁钢高度1
floor_height_flag = True  # 层高
split_no = 1  # 包号格式字母
qty, weight, insert_count = 0, 0, 0
number = 1000 / steel_height
for i in range(len(arr)):
    qty += arr[i][5] * steel_height
    weight += arr[i][7]
    if qty > 1000 or weight > 2500:
        if arr[i][5] * steel_height > 1000:
            def moreno(arr, i_no, qty_no, df_no, insert_count_no):
                density = arr[i_no][7] / arr[i_no][6]  # 密度
                in_column = i_no + insert_count_no
                df_no = pd.concat(
                    [df_no.iloc[:in_column], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                     df_no.iloc[in_column:]], ignore_index=True)
                df_no.iloc[in_column] = df_no.iloc[in_column + 1]
                qty_no_count = qty_no / steel_height
                df_no.iloc[in_column, 5] = number - (qty_no_count - arr[i_no][5])
                df_no.iloc[in_column, 6] = df_no.iloc[in_column, 3] * df_no.iloc[in_column, 4] * df_no.iloc[
                    in_column, 5] / 1000000
                df_no.iloc[in_column, 7] = density * df_no.iloc[in_column, 6]
                arr[i_no][5] -= df_no.iloc[in_column, 5]
                df_no.iloc[in_column + 1, 5] = arr[i_no][5]
                df_no.iloc[in_column + 1, 6] = df_no.iloc[in_column + 1, 3] * df_no.iloc[in_column + 1, 4] * \
                                               df_no.iloc[
                                                   in_column + 1, 5] / 1000000
                df_no.iloc[in_column + 1, 7] = density * df_no.iloc[in_column + 1, 6]
                df_no = pd.concat(
                    [df_no.iloc[:in_column + 1], pd.DataFrame([[None] * len(df_no.columns)], columns=df_no.columns),
                     df_no.iloc[in_column + 1:]], ignore_index=True)
                insert_count_no += 2
                qty_no, weight_no = arr[i_no][5] * steel_height, df_no.iloc[in_column + 1, 7]
                return qty_no, weight_no, insert_count_no, df_no


            while arr[i][5] * steel_height > 1000:
                qty, weight, insert_count, df = moreno(arr, i, qty, df, insert_count)

        else:
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            qty, weight = arr[i][5] * steel_height, arr[i][7]
    elif i > 1:
        if split_string(arr[i][1], split_no) != split_string(arr[i - 1][1], split_no):
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            qty, weight = arr[i][5] * steel_height, arr[i][7]
        elif floor_height_flag == True and arr[i][-1] != arr[i - 1][-1]:
            in_column = i + insert_count
            df = pd.concat([df.iloc[:in_column], pd.DataFrame([[None] * len(df.columns)], columns=df.columns),
                            df.iloc[in_column:]], ignore_index=True)
            insert_count += 1
            qty, weight = arr[i][5] * steel_height, arr[i][7]
arr = np.array(df)
qty, pre_empty, set_beyond = 0, 0, 200
for i in range(len(arr)):
    if not arr[i][1] is None:
        qty += arr[i][5]
    else:
        if qty * steel_height < set_beyond:
            if pre_empty*steel_height < set_beyond:
                df.drop(index=i, inplace=True)
            else:
                df.drop(index=pre_empty, inplace=True)
        qty = 0
        pre_empty = i
pack_no = ''
pack_no_num = 1
pack_no_arr = ['01', '02', '03', '04', '05', '06', '07', '08', '09']
pre_loc = 1
df.loc[len(arr) + 3] = None
arr = np.array(df)
color_arr = []
for i in range(len(arr)):
    if pack_no_num < 10:
        pack_no = pack_no_arr[pack_no_num - 1]
        df.iloc[i, 0] = f'p{pack_no}'
    else:
        df.iloc[i, 0] = f'p{pack_no_num}'
    if arr[i][1] is None or pd.isna(arr[i][1]):
        color_arr.append(i)
        df.iloc[i, 0] = f'小计'
        pack_no_num += 1
        df.iloc[i, 5] = f'=sum(F{pre_loc + 1}:F{i + 1})'
        df.iloc[i, 6] = f'=sum(G{pre_loc + 1}:G{i + 1})'
        df.iloc[i, 7] = f'=sum(H{pre_loc + 1}:H{i + 1})'
        pre_loc = i + 2

df.to_excel("out.xlsx", index=False)

wb = load_workbook("out.xlsx")
ws = wb.active
fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # 黄色

for row_index in color_arr:
    for col_index in range(1, ws.max_column + 1):
        ws.cell(row=row_index + 2, column=col_index).fill = fill

# 保存修改后的工作簿
wb.save("out.xlsx")

os.startfile("out.xlsx")
