# 导入openpyxl库中的load_workbook函数，用于加载Excel工作簿
from openpyxl import load_workbook

# 加载名为"4.xlsx"的Excel工作簿
# wb 是一个 Workbook 对象，代表整个Excel文件
wb = load_workbook("4.xlsx")

# 选择工作簿中名为"8成品二算料单标准"的工作表
# ws 是一个 Worksheet 对象，代表Excel文件中的一个具体工作表
ws = wb["8成品二算料单标准"]

# 获取所有合并单元格区域的列表
# ws.merged_cells.ranges 返回一个包含所有合并区域的迭代器
# list() 将迭代器转换为列表，并创建一个副本。这样做是因为在后续循环中会修改合并单元格，
# 如果不创建副本，直接在迭代器上操作会导致迭代错误。
merged_ranges = list(ws.merged_cells.ranges)

# 遍历每一个合并单元格区域
# merged_range 是一个 CellRange 对象，表示一个合并单元格的范围，例如 'A1:B2'
for merged_range in merged_ranges:
    # 获取合并区域的边界坐标
    # .bounds 属性返回一个元组 (min_col, min_row, max_col, max_row)
    # min_col: 合并区域的最小列索引 (从1开始)
    # min_row: 合并区域的最小行索引 (从1开始)
    # max_col: 合并区域的最大列索引
    # max_row: 合并区域的最大行索引
    min_col, min_row, max_col, max_row = merged_range.bounds

    # 获取合并单元格左上角（即第一个单元格）的值
    # ws.cell(row, column) 用于访问指定行和列的单元格
    top_left_cell_value = ws.cell(row=min_row, column=min_col).value

    # 拆分当前合并单元格
    # ws.unmerge_cells() 方法用于拆分合并单元格
    # str(merged_range) 将 CellRange 对象转换为其字符串表示形式，例如 'A1:B2'
    ws.unmerge_cells(str(merged_range))

    # 填充拆分后的所有单元格
    # 遍历合并区域内的所有行
    for row_idx in range(min_row, max_row + 1):
        # 遍历合并区域内的所有列
        for col_idx in range(min_col, max_col + 1):
            # 将每个拆分后的单元格的值设置为原合并单元格左上角的值
            ws.cell(row=row_idx, column=col_idx).value = top_left_cell_value

# 定义输出文件路径
output_file_path = "4_unmerged.xlsx"

# 获取当前工作簿中所有工作表的名称
# list(wb.sheetnames) 创建一个工作表名称的列表副本，因为在循环中会修改工作表集合
all_sheet_names = list(wb.sheetnames)

# 确定需要保留的工作表的名称
# ws.title 获取当前处理的工作表的名称
sheet_to_keep = ws.title

# 遍历所有工作表名称
for sheet_name in all_sheet_names:
    # 如果工作表名称不是需要保留的，则删除该工作表
    if sheet_name != sheet_to_keep:
        # wb.remove() 方法用于从工作簿中删除指定的工作表
        wb.remove(wb[sheet_name])

# 保存修改后的工作簿到新的Excel文件
# wb.save() 方法用于保存工作簿

wb.save(output_file_path)

# 打印成功消息，显示新文件的保存路径
# f-string (格式化字符串字面量) 用于方便地嵌入表达式
print(f"合并单元格已处理并保存到：{output_file_path}")