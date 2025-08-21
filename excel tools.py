import tkinter as tk
# 导入必要的库
# from tkinter 模块导入 filedialog（文件对话框）、ttk（主题化 Tkinter 部件）
from tkinter import filedialog, ttk
# from tkinter 模块导入 messagebox（消息框）
from tkinter import messagebox
# 导入 pandas 库，用于数据处理，特别是读取和操作 Excel 文件
import pandas as pd
# 导入 numpy 库，用于数值计算，特别是数组操作
import numpy as np
# from openpyxl.workbook 模块导入 Workbook 类，用于创建新的 Excel 工作簿
from openpyxl.workbook import Workbook
# from openpyxl.utils 模块导入 get_column_letter 函数，用于将列索引转换为 Excel 列字母（如 1 -> A, 2 -> B）
from openpyxl.utils import get_column_letter
# 导入 os 模块，用于操作系统相关功能，如文件路径操作
import os
# from tkinterdnd2 模块导入 TkinterDnD（拖放功能主类）和 DND_FILES（拖放文件类型常量）
from tkinterdnd2 import TkinterDnD, DND_FILES

# import time # 这是一个被注释掉的导入，如果需要时间相关功能可以取消注释


# 定义处理 Excel 文件的函数
# file_path 参数是用户选择或拖放的 Excel 文件的路径
def process_excel_file(file_path):
    # 检查文件路径是否存在且不为空
    if file_path:
        try:
            # 使用 pandas 的 read_excel 函数读取指定路径的 Excel 文件，并将其存储为 DataFrame 对象 df
            df = pd.read_excel(file_path)

            # 对 DataFrame 进行转置操作，将行变为列，列变为行
            df_transpose = df.transpose()
            # 获取转置后 DataFrame 的索引（原 DataFrame 的列名），这将作为新 Excel 文件中的标题
            df_index = df_transpose.index

            # 将原始 DataFrame 转换为 NumPy 数组，方便后续按行或列访问数据
            arr = np.array(df)

            # 创建一个新的 Excel 工作簿对象
            wb = Workbook()
            # 获取当前活动的工作表（默认创建时会有一个）
            sheet = wb.active
            # 遍历原始数据的每一行（在转置后的逻辑中，这相当于处理每一列的数据）
            for j in range(0, len(arr)):
                # 设置新 Excel 文件中每隔一列的宽度为 15，用于放置转置后的数据标题
                # get_column_letter(j*2+1) 将列索引转换为字母，例如 j=0 时是 A 列，j=1 时是 C 列
                sheet.column_dimensions[get_column_letter(j*2+1)].width = 15
                # 遍历转置后 DataFrame 的索引（即原始数据的列名）
                for i in range(0, len(df_index)):
                    # 将原始数据的列名写入新 Excel 文件的奇数列（A, C, E...）
                    # sheet.cell(row=i+1, column=j*2+1) 指定单元格位置，i+1 是行号（从1开始），j*2+1 是列号
                    sheet.cell(row=i+1, column=j*2+1).value = df_index[i]
            # 再次遍历原始数据的每一行
            for k in range(0, len(arr)):
                # 设置新 Excel 文件中每隔一列的宽度为 15，用于放置转置后的实际数据
                # get_column_letter(k*2+2) 将列索引转换为字母，例如 k=0 时是 B 列，k=1 时是 D 列
                sheet.column_dimensions[get_column_letter(k*2+2)].width = 15
                # 遍历原始数据的每一列（在转置后的逻辑中，这相当于处理每一行的数据）
                for col_idx in range(0, len(df_index)):
                    # 将原始数据的值写入新 Excel 文件的偶数列（B, D, F...）
                    # arr[k][col_idx] 获取原始 NumPy 数组中对应位置的值
                    sheet.cell(row=col_idx+1, column=k*2+2).value = arr[k][col_idx]

            # 构建输出文件路径
            # 移除原始文件路径中的 .xlsx 或 .xls 扩展名
            file_out = file_path.replace('.xlsx', '')
            file_out = file_out.replace('.xls', '')
            # 保存新的 Excel 工作簿，文件名在原文件基础上加上 '_转置.xlsx'
            wb.save(file_out+'_'+'转置.xlsx')
            # 弹出信息框，显示转置成功消息和新文件的完整路径
            # f-string 用于格式化字符串，方便嵌入变量
            messagebox.showinfo(message=f"转置成功！！\n新文件路径：{file_out}_转置.xlsx")
        # 捕获处理文件过程中可能发生的任何异常
        except Exception as e:
            # 弹出错误信息框，显示错误类型和具体的错误信息
            messagebox.showerror("错误", f"处理文件时发生错误: {e}")


# 定义创建 Excel 转置工具页面的函数
# parent_frame 是这个页面将被放置的父级 Tkinter 框架
def create_transpose_tool_page(parent_frame):
    # 清空父框架中的所有现有组件，以便加载新页面
    for widget in parent_frame.winfo_children():
        widget.destroy()

    # 定义选择文件按钮的回调函数
    def select_file():
        # 弹出文件选择对话框，允许用户选择 Excel 文件
        # title: 对话框的标题
        # initialdir: 初始打开的目录，这里设置为用户的主目录
        # filetypes: 允许选择的文件类型，只显示 Excel 文件和所有文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=os.path.expanduser('~'),
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        # 调用 process_excel_file 函数处理选定的文件
        process_excel_file(file_path)

    # 定义文件拖放事件的回调函数
    # event 对象包含了拖放事件的详细信息，包括文件路径
    def drop_file(event):
        # 从事件数据中获取文件路径
        file_path = event.data
        # 如果文件路径被大括号包围（TkinterDnD 的特性），则移除大括号
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        # 调用 process_excel_file 函数处理拖放的文件
        process_excel_file(file_path)

    # 创建一个标签作为页面标题
    # parent_frame: 标签的父级容器
    # text: 标签显示的文本
    # font: 字体设置，包括字体家族、大小和样式
    # foreground: 文本颜色
    title_label = ttk.Label(parent_frame,
                            text="Excel转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    # 使用 pack 布局管理器放置标签，pady 设置垂直方向的内边距
    title_label.pack(pady=(0, 20))

    # 创建一个提示文本标签
    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))

    # 创建一个 ttk.Style 对象，用于配置 ttk 部件的样式
    style = ttk.Style()
    # 配置 'TButton' 样式，设置按钮的字体和内边距
    style.configure('TButton', font=('微软雅黑', 12), padding=10)

    # 创建一个按钮，用于触发文件选择对话框
    # text: 按钮上显示的文本
    # command: 按钮被点击时调用的函数
    # style: 应用到按钮的样式
    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()

    # 创建一个页脚标签，显示支持的文件格式信息
    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    # 将页脚标签放置在父框架底部，并设置垂直内边距
    footer_label.pack(side='bottom', pady=(20, 0))

    # 注册父框架为拖放目标，允许接收文件拖放
    parent_frame.drop_target_register(DND_FILES)
    # 绑定拖放事件到 drop_file 函数
    parent_frame.dnd_bind('<<Drop>>', drop_file)



# 创建 TkinterDnD 的主窗口实例
root = TkinterDnD.Tk()
# 设置主窗口的标题为 "Excel工具"
root.title("Excel工具")
# 获取屏幕的宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 设置窗口的初始宽度和高度
window_width = 500
window_height = 300

# 计算窗口的 x, y 坐标，使其居中显示
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# 设置窗口的几何尺寸和位置
# 格式为 "宽x高+x坐标+y坐标"
root.geometry(f'{window_width}x{window_height}+{x}+{y}')
# 设置窗口是否可调整大小，True 表示宽度和高度都可调整
root.resizable(True, True)

# 配置主窗口的背景颜色
root.configure(bg='#f0f0f0')

# 创建一个主框架用于容纳所有页面
# root: 父级窗口
# padding: 内边距
main_app_frame = ttk.Frame(root, padding=20)
# 使用 pack 布局管理器，fill='both' 使框架填充可用空间，expand=True 使框架随窗口大小扩展
main_app_frame.pack(fill='both', expand=True)

# 定义显示主页的函数
def show_home_page():
    # 清空主应用框架中的所有现有组件，以便加载主页内容
    for widget in main_app_frame.winfo_children():
        widget.destroy()

    # 创建主页标题标签
    home_title_label = ttk.Label(main_app_frame,
                                 text="Excel工具",
                                 font=('微软雅黑', 18, 'bold'),
                                 foreground='#333333')
    home_title_label.pack(pady=(20, 30))

    # 创建 "Excel 转置工具" 功能按钮
    # text: 按钮文本
    # command: 点击按钮时调用的函数，使用 lambda 表达式延迟调用 create_transpose_tool_page 并传入 main_app_frame
    # style: 应用到按钮的样式
    transpose_btn = ttk.Button(main_app_frame,
                               text="Excel 转置工具",
                               command=lambda: create_transpose_tool_page(main_app_frame),
                               style='TButton')
    transpose_btn.pack(pady=10)

    # 定义创建 Excel 转置工具页面的函数
# parent_frame 是这个页面将被放置的父级 Tkinter 框架
def create_transpose_tool_page(parent_frame):
    # 清空父框架中的所有现有组件，以便加载新页面
    for widget in parent_frame.winfo_children():
        widget.destroy()

    # 定义选择文件按钮的回调函数
    def select_file():
        # 弹出文件选择对话框，允许用户选择 Excel 文件
        # title: 对话框的标题
        # initialdir: 初始打开的目录，这里设置为用户的主目录
        # filetypes: 允许选择的文件类型，只显示 Excel 文件和所有文件
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            initialdir=os.path.expanduser('~'),
            filetypes=[("Excel文件", "*.xls;*.xlsx"), ("所有文件", "*.*")])
        # 调用 process_excel_file 函数处理选定的文件
        process_excel_file(file_path)

    # 定义文件拖放事件的回调函数
    # event 对象包含了拖放事件的详细信息，包括文件路径
    def drop_file(event):
        # 从事件数据中获取文件路径
        file_path = event.data
        # 如果文件路径被大括号包围（TkinterDnD 的特性），则移除大括号
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        # 调用 process_excel_file 函数处理拖放的文件
        process_excel_file(file_path)

    # 创建一个标签作为页面标题
    # parent_frame: 标签的父级容器
    # text: 标签显示的文本
    # font: 字体设置，包括字体家族、大小和样式
    # foreground: 文本颜色
    title_label = ttk.Label(parent_frame,
                            text="Excel转置工具",
                            font=('微软雅黑', 16, 'bold'),
                            foreground='#333333')
    # 使用 pack 布局管理器放置标签，pady 设置垂直方向的内边距
    title_label.pack(pady=(0, 20))

    # 创建一个提示文本标签
    text_label = ttk.Label(parent_frame,
                           text="请选择要转置的Excel文件或将文件拖拽到此处",
                           font=('微软雅黑', 12))
    text_label.pack(pady=(0, 30))

    # 创建一个 ttk.Style 对象，用于配置 ttk 部件的样式
    style = ttk.Style()
    # 配置 'TButton' 样式，设置按钮的字体和内边距
    style.configure('TButton', font=('微软雅黑', 12), padding=10)

    # 创建一个按钮，用于触发文件选择对话框
    # text: 按钮上显示的文本
    # command: 按钮被点击时调用的函数
    # style: 应用到按钮的样式
    btn = ttk.Button(parent_frame,
                     text="选择文件",
                     command=select_file,
                     style='TButton')
    btn.pack()

    # 创建一个页脚标签，显示支持的文件格式信息
    footer_label = ttk.Label(parent_frame,
                             text="支持.xls和.xlsx格式",
                             font=('微软雅黑', 9),
                             foreground='#666666')
    # 将页脚标签放置在父框架底部，并设置垂直内边距
    footer_label.pack(side='bottom', pady=(20, 0))

    # 注册父框架为拖放目标，允许接收文件拖放
    parent_frame.drop_target_register(DND_FILES)
    # 绑定拖放事件到 drop_file 函数
    parent_frame.dnd_bind('<<Drop>>', drop_file)

    # 添加返回主页按钮
    back_button = ttk.Button(parent_frame,
                             text="返回主页",
                             command=show_home_page,
                             style='TButton')
    back_button.pack(pady=10)

# 应用程序启动时，初始显示主页
show_home_page()

# 启动 Tkinter 事件循环，使 GUI 应用程序运行
root.mainloop()
